# ============================================================
# Nhập lịch làm việc (Thứ 7 / Chủ nhật) bằng Excel — file mẫu + parser.
# Owner: Nhật Anh.
#
# Tách khỏi controllers/main.py để test thẳng bằng hàm thuần: sinh file mẫu →
# parse lại → kiểm từng nhánh lỗi, không cần dựng HTTP request.
#
# Quy ước file mẫu: 1 sheet, 3 dòng đầu là tiêu đề + hướng dẫn, dòng thứ 4 là
# header, mỗi dòng sau là MỘT ngày Thứ 7 / Chủ nhật CHƯA ĐẾN của năm đó. HR chỉ
# điền 'x' vào cột "Đi làm" (2 cột Ngày/Thứ bị khoá bằng sheet protection) —
# nhờ vậy không thể chọn nhầm ngày trong tuần hay ngày của năm khác.
# ============================================================
import io
from datetime import date, datetime, timedelta

# weekday() 0..6 → nhãn tiếng Việt.
DOW_VN = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6', 'Thứ 7', 'Chủ nhật']

COL_DATE = 'Ngày (dd/mm/yyyy)'
COL_DOW = 'Thứ'
COL_MARK = 'Đi làm (x)'
COL_NOTE = 'Ghi chú'
HEADERS = [COL_DATE, COL_DOW, COL_MARK, COL_NOTE]

# Giá trị được coi là "có đi làm" ở cột Đi làm. Nhận rộng tay vì HR gõ tự do,
# nhưng giá trị LẠ thì báo lỗi chứ không bỏ qua im lặng (dễ mất ngày).
MARK_YES = frozenset({'x', 'v', '1', 'có', 'co', 'yes', 'true', 'x.'})
MARK_NO = frozenset({'', '-', '0', 'không', 'khong', 'no', 'false'})

DEFAULT_NAME = 'Ngày làm bù'
MAX_ERRORS = 15          # chỉ liệt kê ngần này dòng lỗi cho dễ đọc
MAX_XLSX_BYTES = 2 * 1024 * 1024
XLSX_MIME = ('application/vnd.openxmlformats-officedocument'
             '.spreadsheetml.sheet')


class WorkdayImportError(Exception):
    """Lỗi định dạng/nội dung file import — controller đổi thẳng ra JSON 400."""

    def __init__(self, code, message, details=None):
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = details or []


# ----------------------------------------------------------------- helpers
def weekend_days(year, min_date=None):
    """Mọi ngày Thứ 7 / Chủ nhật của `year`, bỏ ngày < min_date (đã đến)."""
    out, cur = [], date(year, 1, 1)
    while cur.year == year:
        if cur.weekday() >= 5 and (not min_date or cur >= min_date):
            out.append(cur)
        cur += timedelta(days=1)
    return out


def _norm(val):
    """Ô Excel → chuỗi thường, đã trim (None → '')."""
    if val is None:
        return ''
    if isinstance(val, str):
        return val.strip().lower()
    return str(val).strip().lower()


def _to_date(val):
    """Ô Excel → date. Nhận cả ô kiểu ngày lẫn chuỗi dd/mm/yyyy | yyyy-mm-dd."""
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        txt = val.strip()
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
            try:
                return datetime.strptime(txt, fmt).date()
            except ValueError:
                continue
    return None


def _fmt(day):
    return day.strftime('%d/%m/%Y')


# ------------------------------------------------------------- file mẫu
def build_template(year, min_date=None):
    """Sinh file .xlsx mẫu (bytes) liệt kê Thứ 7/Chủ nhật chưa đến của `year`.

    Không bao giờ raise khi năm hết ngày hợp lệ — vẫn trả file (có dòng ghi
    chú) vì nút tải mẫu là link tải thẳng, trả JSON lỗi sẽ hiện raw ra tab mới.
    """
    try:
        import xlsxwriter
    except ImportError:
        raise WorkdayImportError(
            'no_xlsxwriter',
            'Server chưa cài thư viện xlsxwriter nên không tạo được file mẫu.')

    days = weekend_days(year, min_date)
    out = io.BytesIO()
    wb = xlsxwriter.Workbook(out, {'in_memory': True})
    ws = wb.add_worksheet('Lịch làm việc %d' % year)

    f_title = wb.add_format({'bold': True, 'font_size': 13, 'font_color': '#7A0C1E'})
    f_hint = wb.add_format({'font_size': 10, 'italic': True, 'font_color': '#78716C',
                            'text_wrap': True, 'valign': 'top'})
    f_head = wb.add_format({'bold': True, 'bg_color': '#F8DADF', 'border': 1,
                            'align': 'center', 'valign': 'vcenter'})
    f_lock = wb.add_format({'border': 1, 'bg_color': '#FAF8F5', 'locked': 1})
    f_lock_c = wb.add_format({'border': 1, 'bg_color': '#FAF8F5', 'locked': 1,
                              'align': 'center'})
    f_open = wb.add_format({'border': 1, 'locked': 0, 'align': 'center'})
    f_open_l = wb.add_format({'border': 1, 'locked': 0})

    ws.set_column(0, 0, 18)
    ws.set_column(1, 1, 12)
    ws.set_column(2, 2, 12)
    ws.set_column(3, 3, 34)

    ws.write(0, 0, 'LỊCH LÀM VIỆC THỨ 7 / CHỦ NHẬT — NĂM %d' % year, f_title)
    ws.merge_range(1, 0, 1, 3, (
        'Cách dùng: điền "x" vào cột "Đi làm (x)" ở những ngày cần đi làm bù, '
        'ghi chú (nếu có) ở cột cuối, rồi tải file này lên lại. '
        'Hai cột Ngày và Thứ đã khoá — không sửa, không thêm dòng. '
        'File chỉ liệt kê Thứ 7 / Chủ nhật CHƯA ĐẾN của năm %d.' % year), f_hint)
    ws.set_row(1, 34)

    head_row = 3
    for col, label in enumerate(HEADERS):
        ws.write(head_row, col, label, f_head)

    for i, day in enumerate(days):
        r = head_row + 1 + i
        ws.write_string(r, 0, _fmt(day), f_lock)
        ws.write_string(r, 1, DOW_VN[day.weekday()], f_lock_c)
        ws.write_blank(r, 2, None, f_open)
        ws.write_blank(r, 3, None, f_open_l)

    if days:
        first, last = head_row + 1, head_row + len(days)
        ws.data_validation(first, 2, last, 2, {
            'validate': 'list',
            'source': ['x'],
            'input_title': 'Đi làm?',
            'input_message': 'Điền "x" nếu ngày này đi làm bù, để trống nếu nghỉ.',
            'error_title': 'Giá trị không hợp lệ',
            'error_message': 'Chỉ điền "x" hoặc để trống.',
        })
    else:
        ws.merge_range(head_row + 1, 0, head_row + 1, 3,
                       'Năm %d không còn ngày Thứ 7 / Chủ nhật nào chưa đến — '
                       'không thêm được lịch làm việc cho ngày đã qua.' % year,
                       f_hint)

    ws.freeze_panes(head_row + 1, 0)
    # Khoá sheet (không mật khẩu): chỉ 2 cột "Đi làm"/"Ghi chú" còn sửa được.
    ws.protect('', {'select_locked_cells': True, 'select_unlocked_cells': True})
    wb.close()
    return out.getvalue()


# --------------------------------------------------------------- parser
def _locate_header(ws):
    """Tìm dòng header + vị trí các cột. Trả (row_idx, {key: col_idx})."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=12,
                                               values_only=True), start=1):
        cells = [_norm(c) for c in (row or [])]
        cols = {}
        for i, txt in enumerate(cells):
            if not txt:
                continue
            if 'ngày' in txt and 'date' not in cols:
                cols['date'] = i
            elif txt.startswith('thứ') and 'dow' not in cols:
                cols['dow'] = i
            elif 'đi làm' in txt and 'mark' not in cols:
                cols['mark'] = i
            elif 'ghi chú' in txt and 'note' not in cols:
                cols['note'] = i
        if 'date' in cols and 'mark' in cols:
            return row_idx, cols
    raise WorkdayImportError(
        'bad_template',
        'File không đúng mẫu: không tìm thấy dòng tiêu đề có cột "%s" và "%s". '
        'Hãy tải lại file mẫu ở trên và điền vào đúng file đó.'
        % (COL_DATE, COL_MARK))


def parse_workdays_xlsx(content, year, min_date=None, existing=None):
    """Đọc file .xlsx HR tải lên → danh sách ngày đi làm đã được kiểm.

    Trả {'rows': [{date, dow, name}], 'skipped': [{date, dow, reason}]}.
    Mọi sai định dạng đều raise WorkdayImportError NGAY (không nhập một phần).
    """
    try:
        import openpyxl
    except ImportError:
        raise WorkdayImportError(
            'no_openpyxl',
            'Server chưa cài thư viện openpyxl nên không đọc được file Excel.')

    if not content:
        raise WorkdayImportError('empty_file', 'File rỗng, chưa có nội dung.')
    if len(content) > MAX_XLSX_BYTES:
        raise WorkdayImportError(
            'too_large', 'File quá lớn (tối đa %d MB).'
            % (MAX_XLSX_BYTES // (1024 * 1024)))

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True,
                                    read_only=True)
    except Exception:
        raise WorkdayImportError(
            'bad_file',
            'Không đọc được file: file hỏng hoặc không phải Excel .xlsx. '
            'Hãy tải lại file mẫu và điền vào đúng file đó.')

    ws = wb.worksheets[0] if wb.worksheets else None
    if ws is None:
        raise WorkdayImportError('bad_file', 'File Excel không có sheet nào.')

    head_row, cols = _locate_header(ws)
    c_date, c_mark = cols['date'], cols['mark']
    c_note = cols.get('note')

    existing = set(existing or ())
    errors, rows, skipped, seen = [], [], [], {}

    for r_idx, row in enumerate(ws.iter_rows(min_row=head_row + 1,
                                             values_only=True),
                                start=head_row + 1):
        row = row or ()
        if not any(c not in (None, '') for c in row):
            continue

        raw_mark = row[c_mark] if c_mark < len(row) else None
        mark = _norm(raw_mark)
        if mark in MARK_NO:
            continue
        if mark not in MARK_YES:
            errors.append('Dòng %d: cột "%s" chỉ nhận "x" (hoặc để trống), '
                          'đang là "%s".' % (r_idx, COL_MARK, raw_mark))
            continue

        day = _to_date(row[c_date] if c_date < len(row) else None)
        if not day:
            errors.append('Dòng %d: cột "%s" không đọc được ngày (cần dạng '
                          'dd/mm/yyyy).' % (r_idx, COL_DATE))
            continue
        if day.year != year:
            errors.append('Dòng %d: ngày %s không thuộc năm %d.'
                          % (r_idx, _fmt(day), year))
            continue
        if day.weekday() < 5:
            errors.append('Dòng %d: ngày %s là %s — file này chỉ nhận Thứ 7 và '
                          'Chủ nhật.' % (r_idx, _fmt(day), DOW_VN[day.weekday()]))
            continue
        if min_date and day < min_date:
            errors.append('Dòng %d: ngày %s đã đến hoặc đã qua — chỉ thêm được '
                          'từ ngày %s trở đi.' % (r_idx, _fmt(day), _fmt(min_date)))
            continue
        if day in seen:
            errors.append('Dòng %d: ngày %s bị lặp (đã có ở dòng %d).'
                          % (r_idx, _fmt(day), seen[day]))
            continue

        seen[day] = r_idx
        dow = DOW_VN[day.weekday()]
        if day in existing:
            skipped.append({'date': day.isoformat(), 'dow': dow,
                            'reason': 'Đã có trong lịch'})
            continue
        note = row[c_note] if (c_note is not None and c_note < len(row)) else None
        name = (str(note).strip() if note not in (None, '') else '') or DEFAULT_NAME
        rows.append({'date': day.isoformat(), 'dow': dow, 'name': name})

    wb.close()

    if errors:
        raise WorkdayImportError(
            'bad_rows',
            'File có %d dòng không hợp lệ — chưa nhập ngày nào. Sửa lại rồi '
            'tải lên lần nữa.' % len(errors),
            details=errors[:MAX_ERRORS] + (
                ['… và %d dòng lỗi khác.' % (len(errors) - MAX_ERRORS)]
                if len(errors) > MAX_ERRORS else []))
    if not rows and not skipped:
        raise WorkdayImportError(
            'empty_selection',
            'Chưa đánh dấu ngày nào: điền "x" vào cột "%s" ở những ngày cần đi '
            'làm bù rồi tải lên lại.' % COL_MARK)

    return {'rows': rows, 'skipped': skipped}
