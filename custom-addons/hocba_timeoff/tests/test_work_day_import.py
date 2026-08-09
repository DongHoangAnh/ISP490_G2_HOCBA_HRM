"""Nhập lịch làm việc Thứ 7 / Chủ nhật từ Excel (controllers/workday_xlsx.py).

Kiểm hàm thuần: sinh file mẫu → điền → parse lại. Trọng tâm là các nhánh SAI
định dạng phải báo lỗi NGAY và KHÔNG nhập một phần (raise trước khi trả rows).
"""
import io
from datetime import date, timedelta

import openpyxl
from odoo import fields
from odoo.tests import TransactionCase, tagged

from odoo.addons.hocba_timeoff.controllers.workday_xlsx import (
    COL_MARK, DEFAULT_NAME, WorkdayImportError, build_template,
    parse_workdays_xlsx, weekend_days,
)


@tagged('post_install', '-at_install', 'hocba_timeoff')
class TestWorkDayImport(TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.today = fields.Date.context_today(cls.env['hb.work.day'])
        cls.min_date = cls.today + timedelta(days=1)
        # Năm sau: mọi ngày đều > min_date → test không phụ thuộc ngày chạy.
        cls.year = cls.today.year + 1
        cls.days = weekend_days(cls.year)

    # ---------------------------------------------------------------- mẫu
    def _template(self, year=None, min_date=None):
        return build_template(year or self.year, min_date)

    def _fill(self, content, marks, notes=None):
        """Điền cột "Đi làm" cho file mẫu. marks: {ngày dd/mm/yyyy → giá trị}."""
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.worksheets[0]
        for row in ws.iter_rows(min_row=5):
            key = row[0].value
            if key in marks:
                row[2].value = marks[key]
                if notes and key in notes:
                    row[3].value = notes[key]
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    def _parse(self, content, year=None, existing=None):
        return parse_workdays_xlsx(content, year or self.year,
                                   self.min_date, existing)

    def test_template_lists_only_weekends(self):
        wb = openpyxl.load_workbook(io.BytesIO(self._template()))
        ws = wb.worksheets[0]
        rows = [r[0].value for r in ws.iter_rows(min_row=5) if r[0].value]
        self.assertEqual(len(rows), len(self.days))
        for txt in rows:
            d, m, y = (int(x) for x in txt.split('/'))
            self.assertEqual(y, self.year)
            self.assertGreaterEqual(date(y, m, d).weekday(), 5,
                                    '%s không phải Thứ 7/Chủ nhật' % txt)

    def test_template_skips_past_days(self):
        """Năm hiện tại: file mẫu chỉ liệt kê ngày CHƯA ĐẾN."""
        wb = openpyxl.load_workbook(
            io.BytesIO(build_template(self.today.year, self.min_date)))
        ws = wb.worksheets[0]
        for r in ws.iter_rows(min_row=5):
            if not r[0].value or '/' not in str(r[0].value):
                continue
            d, m, y = (int(x) for x in r[0].value.split('/'))
            self.assertGreaterEqual(date(y, m, d), self.min_date)

    # -------------------------------------------------------------- happy
    def test_parse_marked_days(self):
        picked = [self.days[0], self.days[3]]
        content = self._fill(
            self._template(),
            {d.strftime('%d/%m/%Y'): 'x' for d in picked},
            {picked[0].strftime('%d/%m/%Y'): 'Làm bù Tết'})
        res = self._parse(content)
        self.assertEqual([r['date'] for r in res['rows']],
                         [d.isoformat() for d in picked])
        self.assertEqual(res['rows'][0]['name'], 'Làm bù Tết')
        self.assertEqual(res['rows'][1]['name'], DEFAULT_NAME)
        self.assertEqual(res['skipped'], [])

    def test_mark_variants_accepted(self):
        content = self._fill(self._template(), {
            self.days[0].strftime('%d/%m/%Y'): 'X',
            self.days[1].strftime('%d/%m/%Y'): ' có ',
        })
        self.assertEqual(len(self._parse(content)['rows']), 2)

    def test_existing_day_skipped_not_error(self):
        picked = [self.days[0], self.days[1]]
        content = self._fill(self._template(),
                             {d.strftime('%d/%m/%Y'): 'x' for d in picked})
        res = self._parse(content, existing={picked[0]})
        self.assertEqual([r['date'] for r in res['rows']],
                         [picked[1].isoformat()])
        self.assertEqual(len(res['skipped']), 1)

    # ------------------------------------------------------------- lỗi
    def test_not_xlsx_content(self):
        with self.assertRaises(WorkdayImportError) as cm:
            self._parse(b'day khong phai file excel')
        self.assertEqual(cm.exception.code, 'bad_file')

    def test_empty_file(self):
        with self.assertRaises(WorkdayImportError) as cm:
            self._parse(b'')
        self.assertEqual(cm.exception.code, 'empty_file')

    def test_wrong_template_headers(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Cột lạ', 'Cột khác'])
        ws.append(['01/01/2030', 'x'])
        out = io.BytesIO()
        wb.save(out)
        with self.assertRaises(WorkdayImportError) as cm:
            self._parse(out.getvalue())
        self.assertEqual(cm.exception.code, 'bad_template')

    def test_nothing_marked(self):
        with self.assertRaises(WorkdayImportError) as cm:
            self._parse(self._template())
        self.assertEqual(cm.exception.code, 'empty_selection')

    def test_unknown_mark_value_is_error(self):
        """Giá trị lạ ở cột Đi làm KHÔNG được bỏ qua im lặng — dễ mất ngày."""
        content = self._fill(self._template(),
                             {self.days[0].strftime('%d/%m/%Y'): 'lam bu'})
        with self.assertRaises(WorkdayImportError) as cm:
            self._parse(content)
        self.assertEqual(cm.exception.code, 'bad_rows')
        self.assertIn(COL_MARK, cm.exception.details[0])

    def _tampered(self, new_date, mark='x', row=5):
        """Sửa tay ô ngày (mô phỏng HR gỡ khoá sheet rồi gõ đè)."""
        wb = openpyxl.load_workbook(io.BytesIO(self._template()))
        ws = wb.worksheets[0]
        ws.cell(row=row, column=1).value = new_date
        ws.cell(row=row, column=3).value = mark
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    def test_weekday_rejected(self):
        monday = next(d for d in weekend_days(self.year)
                      if True) + timedelta(days=2)   # Thứ 7 + 2 = Thứ 2
        with self.assertRaises(WorkdayImportError) as cm:
            self._parse(self._tampered(monday.strftime('%d/%m/%Y')))
        self.assertEqual(cm.exception.code, 'bad_rows')
        self.assertIn('Thứ 2', cm.exception.details[0])

    def test_other_year_rejected(self):
        other = weekend_days(self.year + 1)[0]
        with self.assertRaises(WorkdayImportError) as cm:
            self._parse(self._tampered(other.strftime('%d/%m/%Y')))
        self.assertEqual(cm.exception.code, 'bad_rows')
        self.assertIn('không thuộc năm', cm.exception.details[0])

    def test_past_day_rejected(self):
        past = self.today - timedelta(days=7)
        while past.weekday() < 5:
            past -= timedelta(days=1)
        with self.assertRaises(WorkdayImportError) as cm:
            parse_workdays_xlsx(self._tampered(past.strftime('%d/%m/%Y')),
                                past.year, self.min_date, None)
        self.assertEqual(cm.exception.code, 'bad_rows')

    def test_unparsable_date_rejected(self):
        with self.assertRaises(WorkdayImportError) as cm:
            self._parse(self._tampered('ngày mai'))
        self.assertEqual(cm.exception.code, 'bad_rows')

    def test_duplicate_date_rejected(self):
        dup = self.days[1].strftime('%d/%m/%Y')
        content = self._tampered(dup, row=5)      # dòng 5 = dòng của days[0]
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.worksheets[0]
        ws.cell(row=6, column=3).value = 'x'      # dòng 6 = days[1] gốc
        out = io.BytesIO()
        wb.save(out)
        with self.assertRaises(WorkdayImportError) as cm:
            self._parse(out.getvalue())
        self.assertEqual(cm.exception.code, 'bad_rows')
        self.assertIn('lặp', cm.exception.details[0])

    def test_no_partial_import_on_error(self):
        """1 dòng sai → cả lô bị chặn, không trả về ngày hợp lệ nào."""
        good = self.days[0].strftime('%d/%m/%Y')
        content = self._fill(self._tampered('sai bét'), {good: 'x'})
        with self.assertRaises(WorkdayImportError):
            self._parse(content)
