"""Nhập hồ sơ nhân viên từ Excel (controllers/employee_xlsx.py + employee_import.py).

Trọng tâm: parser là hàm thuần nên test thẳng, không dựng HTTP. Các nhánh danh
mục SAI phải báo lỗi kèm số dòng Excel chứ không âm thầm rơi về mặc định.
"""
import io

import openpyxl
from odoo.tests.common import TransactionCase
from odoo.tests import tagged

from odoo.addons.hocba_hrm.controllers.employee_xlsx import (
    EmployeeImportError, HEADER_ALIASES, locate_header, norm_header,
)


def _sheet(rows):
    """list[list] → worksheet openpyxl (dòng 1 của file = rows[0])."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    out = io.BytesIO()
    wb.save(out)
    return openpyxl.load_workbook(io.BytesIO(out.getvalue())).worksheets[0]


HDR = ['Mã nhân sự', 'Họ và tên', 'Tình trạng', 'Phòng Ban', 'Chức danh']


@tagged('post_install', '-at_install')
class TestEmployeeXlsxHeader(TransactionCase):

    def test_norm_header_stripped_accents_and_suffix(self):
        self.assertEqual(norm_header('Họ và tên'), 'ho va ten')
        self.assertEqual(norm_header('Họ tên nhân sự_text'), 'ho ten nhan su')
        self.assertEqual(norm_header('  Phòng  Ban '), 'phong ban')
        self.assertEqual(norm_header(None), '')

    def test_locate_header_on_first_row(self):
        ws = _sheet([HDR, ['HB.01', 'Nguyễn A', 'Chính thức', 'Marketing', 'NV']])
        row, colmap, unknown = locate_header(ws)
        self.assertEqual(row, 1)
        self.assertEqual(colmap['code'], 0)
        self.assertEqual(colmap['name'], 1)
        self.assertEqual(colmap['status'], 2)
        self.assertEqual(unknown, [])

    def test_locate_header_below_title_rows(self):
        ws = _sheet([['DANH SÁCH NHÂN SỰ'], [], HDR,
                     ['HB.01', 'Nguyễn A', 'Chính thức', 'Marketing', 'NV']])
        row, colmap, _u = locate_header(ws)
        self.assertEqual(row, 3)
        self.assertEqual(colmap['name'], 1)

    def test_locate_header_missing_raises(self):
        ws = _sheet([['a', 'b'], ['c', 'd']])
        with self.assertRaises(EmployeeImportError) as cm:
            locate_header(ws)
        self.assertEqual(cm.exception.code, 'no_header')

    def test_locate_header_without_name_column_raises(self):
        ws = _sheet([['Mã nhân sự', 'Tình trạng', 'Phòng Ban', 'Chức danh']])
        with self.assertRaises(EmployeeImportError) as cm:
            locate_header(ws)
        self.assertEqual(cm.exception.code, 'no_name_col')

    def test_unknown_columns_collected(self):
        ws = _sheet([HDR + ['Tài khoản Lark', 'Màn máy tính']])
        _row, colmap, unknown = locate_header(ws)
        self.assertNotIn('lark', colmap)
        self.assertEqual(unknown, ['Tài khoản Lark', 'Màn máy tính'])

    def test_company_email_wins_over_personal(self):
        ws = _sheet([['Họ và tên', 'Email cá nhân', 'Email công ty']])
        _row, colmap, _u = locate_header(ws)
        self.assertEqual(colmap['email'], 2)

    def test_personal_email_used_when_no_company_column(self):
        ws = _sheet([['Họ và tên', 'Email cá nhân', 'Số điện thoại']])
        _row, colmap, _u = locate_header(ws)
        self.assertEqual(colmap['email'], 1)

    def test_every_alias_maps_to_known_key(self):
        keys = {'code', 'name', 'status', 'workForm', 'posType', 'dep', 'job',
                'probStart', 'bday', 'cccd', 'idIssue', 'idPlace', 'phone',
                'email', 'emailAlt', 'bankAccountNo', 'bankCode', 'pit', 'si',
                'hi', 'hiPlace', 'permStreet', 'currStreet'}
        self.assertEqual(set(HEADER_ALIASES.values()) - keys, set())


from odoo.addons.hocba_hrm.controllers.employee_xlsx import parse_employees_xlsx


def _book(rows, sheet_title='2.1. Quản lý nhân sự'):
    """list[list] → bytes .xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    for r in rows:
        ws.append(r)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


FULL_HDR = ['Mã nhân sự', 'Họ và tên', 'Tình trạng', 'Hình thức', 'Loại vị trí',
            'Phòng Ban', 'Chức danh', 'Ngày thử việc', 'Số căn cước công dân',
            'Số điện thoại', 'Email công ty', 'Mã số thuế TNCN', 'Số sổ BHXH']


def _row(code='HB.01', name='Nguyễn Văn A', status='Chính thức', form='Offline',
         pos='Nhân viên', dep='Marketing', job='Nhân viên R&D', prob='01/06/2024',
         cccd='038098029187', phone='0325252626', email='a@hocba.vn',
         pit=None, si=None):
    return [code, name, status, form, pos, dep, job, prob, cccd, phone, email, pit, si]


# Khoá dựng y hệt cách controller làm (_catalogs dùng norm_header trên tên
# thật) — đặt khoá thô ở đây là fixture nói dối, che mất lỗi ánh xạ bí danh.
CATALOGS = {
    'departments': {norm_header('Marketing'): 11,
                    norm_header('Sản phẩm (R&D_SP)'): 12,
                    norm_header('Kế toán_HCNS'): 13},
    'jobs': {norm_header('Nhân viên R&D'): 21},
}


@tagged('post_install', '-at_install')
class TestEmployeeXlsxParse(TransactionCase):

    def _parse(self, rows, existing=None, hdr=None):
        return parse_employees_xlsx(_book([hdr or FULL_HDR] + rows),
                                    catalogs=CATALOGS, existing=existing)

    def test_happy_row(self):
        res = self._parse([_row()])
        self.assertEqual(res['summary']['ok'], 1)
        r = res['rows'][0]
        self.assertEqual(r['excelRow'], 2)
        self.assertEqual(r['values']['name'], 'Nguyễn Văn A')
        self.assertEqual(r['values']['x_employment_status'], 'official')
        self.assertEqual(r['values']['x_work_form'], 'offline')
        self.assertEqual(r['values']['x_position_type'], 'staff')
        self.assertEqual(r['values']['department_id'], 11)
        self.assertEqual(r['values']['job_id'], 21)
        self.assertEqual(r['values']['cccd'], '038098029187')

    def test_cccd_keeps_leading_zero(self):
        res = self._parse([_row(cccd='038098029187')])
        self.assertEqual(res['rows'][0]['values']['cccd'], '038098029187')

    def test_cccd_wrong_length_warns_and_blanks(self):
        res = self._parse([_row(cccd='12345')])
        r = res['rows'][0]
        self.assertNotIn('cccd', r['values'])
        self.assertTrue(any('CCCD' in w for w in r['warnings']))

    def test_status_online_is_row_error(self):
        res = self._parse([_row(status='Online')])
        self.assertEqual(res['rows'], [])
        self.assertEqual(res['errors'][0]['code'], 'bad_status')
        self.assertEqual(res['errors'][0]['excelRow'], 2)

    def test_department_alias_matched(self):
        res = self._parse([_row(dep='Phòng R&D_SP')])
        self.assertEqual(res['errors'], [])
        self.assertEqual(res['rows'][0]['values']['department_id'], 12)

    def test_department_alias_ke_toan(self):
        res = self._parse([_row(dep='Kế Toán')])
        self.assertEqual(res['rows'][0]['values']['department_id'], 13)

    def test_department_matched_case_insensitively(self):
        res = self._parse([_row(dep='kinh doanh')])
        self.assertEqual(res['errors'][0]['code'], 'bad_department')
        res = self._parse([_row(dep='MARKETING')])
        self.assertEqual(res['rows'][0]['values']['department_id'], 11)

    def test_unknown_department_is_row_error(self):
        res = self._parse([_row(dep='Phòng ma')])
        self.assertEqual(res['errors'][0]['code'], 'bad_department')

    def test_unknown_job_kept_as_text(self):
        res = self._parse([_row(job='Nhân viên tư vấn tuyển sinh thử việc')])
        vals = res['rows'][0]['values']
        self.assertNotIn('job_id', vals)
        self.assertEqual(vals['job_title'], 'Nhân viên tư vấn tuyển sinh thử việc')

    def test_dates_accept_three_string_formats(self):
        for txt in ('01/06/2024', '2024-06-01', '01-06-2024'):
            res = self._parse([_row(prob=txt)])
            self.assertEqual(str(res['rows'][0]['values']['x_probation_start']),
                             '2024-06-01', txt)

    def test_bad_date_warns_and_blanks(self):
        res = self._parse([_row(prob='Cần ngay')])
        r = res['rows'][0]
        self.assertNotIn('x_probation_start', r['values'])
        self.assertTrue(r['warnings'])

    def test_bad_email_warns_and_blanks(self):
        res = self._parse([_row(email='khong-phai-email')])
        self.assertNotIn('work_email', res['rows'][0]['values'])

    def test_empty_name_is_row_error(self):
        res = self._parse([_row(name='   ')])
        self.assertEqual(res['errors'][0]['code'], 'empty_name')

    def test_blank_rows_ignored_not_errors(self):
        res = self._parse([_row(), [None] * len(FULL_HDR)])
        self.assertEqual(res['summary']['total'], 1)
        self.assertEqual(res['errors'], [])

    def test_existing_code_skipped(self):
        res = self._parse([_row(code='HB.01')],
                          existing={'codes': {'HB.01'}, 'cccds': set()})
        self.assertEqual(res['rows'], [])
        self.assertEqual(res['skipped'][0]['reason'], 'code_exists')

    def test_existing_cccd_skipped(self):
        res = self._parse([_row(code='HB.99')],
                          existing={'codes': set(), 'cccds': {'038098029187'}})
        self.assertEqual(res['skipped'][0]['reason'], 'cccd_exists')

    def test_duplicate_code_inside_file_skipped_second(self):
        res = self._parse([_row(code='HB.07'), _row(code='HB.07', cccd='038098029188')])
        self.assertEqual(len(res['rows']), 1)
        self.assertEqual(res['skipped'][0]['reason'], 'code_exists')

    def test_missing_official_fields_reported(self):
        res = self._parse([_row(pit=None, si=None)])
        self.assertEqual(res['rows'][0]['missingOfficial'], ['MST TNCN', 'Số sổ BHXH'])
        self.assertEqual(res['summary']['needCompletion'], 1)

    def test_probation_row_has_no_missing_official(self):
        res = self._parse([_row(status='Thử việc', pit=None, si=None)])
        self.assertEqual(res['rows'][0]['missingOfficial'], [])

    def test_sheet_list_and_chosen_sheet(self):
        res = self._parse([_row()])
        self.assertEqual(res['sheet'], '2.1. Quản lý nhân sự')
        self.assertIn('2.1. Quản lý nhân sự', res['sheets'])


from odoo.exceptions import ValidationError


@tagged('post_install', '-at_install')
class TestLegacyImportEscapeHatch(TransactionCase):
    """BR-010 phải MỞ cho luồng nhập dữ liệu cũ và ĐÓNG cho mọi luồng khác.

    Lý do miễn: 29/112 nhân sự cũ của trung tâm đang là "Chính thức" ngoài đời
    nhưng chưa từng đi qua quy trình thử việc của hệ thống (cả file không ai có
    MST). Nhập trọn gói nên chỉ cần một dòng vướng là cả mẻ bị huỷ.
    """

    VALS = {'name': 'NV Cũ', 'x_employment_status': 'official'}

    def test_official_without_pit_still_blocked_normally(self):
        with self.assertRaises(ValidationError):
            self.env['hr.employee'].create(dict(self.VALS))

    def test_official_without_pit_allowed_under_import_context(self):
        emp = self.env['hr.employee'].with_context(
            hocba_legacy_import=True).create(dict(self.VALS))
        self.assertEqual(emp.x_employment_status, 'official')
        self.assertEqual(
            set(emp._hocba_missing_official_fields()),
            {'CCCD', 'MST TNCN', 'Số sổ BHXH'})

    def test_promoting_imported_record_later_is_still_blocked(self):
        """Hồ sơ đã nhập KHÔNG được tha luôn: đường lên chính thức bình thường
        vẫn phải khai đủ."""
        imported = self.env['hr.employee'].with_context(
            hocba_legacy_import=True).create(dict(self.VALS))
        # Cờ context DÍNH vào recordset vừa tạo. HR sửa hồ sơ ở một request
        # khác, không có cờ — browse lại để mô phỏng đúng chuyện đó.
        emp = self.env['hr.employee'].browse(imported.id)
        self.assertNotIn('hocba_legacy_import', emp.env.context)
        emp.write({'x_employment_status': 'probation'})
        with self.assertRaises(ValidationError):
            emp.write({'x_employment_status': 'official'})

    def test_clearing_pit_on_official_still_blocked(self):
        # CCCD nằm trên hr.version nên phải ghi TRƯỚC khi lên official, không
        # thì vướng BR-010 ngay lúc tạo (chính là bẫy ghi trong CLAUDE.md).
        emp = self.env['hr.employee'].create({
            'name': 'NV Đủ giấy tờ', 'x_employment_status': 'probation',
            'x_pit_code': '0000000001', 'x_social_insurance_no': '1900000000'})
        emp.version_id.identification_id = '012345678901'
        emp.write({'x_employment_status': 'official'})
        with self.assertRaises(ValidationError):
            emp.write({'x_pit_code': False})

    def test_probation_never_needed_the_papers(self):
        emp = self.env['hr.employee'].create({
            'name': 'NV Thử việc', 'x_employment_status': 'probation'})
        self.assertEqual(emp.x_employment_status, 'probation')


from odoo import http

from odoo.addons.hocba_hrm.controllers.employee_import import HocBaEmployeeImport
from odoo.addons.hocba_hrm.controllers.main import _cap_import_emp


class _FakeUpload:
    """Stand-in cho werkzeug FileStorage: đủ .filename và .read()."""

    def __init__(self, content, filename='ns.xlsx'):
        self._content = content
        self.filename = filename

    def read(self):
        return self._content


class _FakeRequest:
    """Đủ để controller đọc request.env khi gọi thẳng ngoài context HTTP."""

    def __init__(self, env):
        self.env = env


@tagged('post_install', '-at_install')
class TestEmployeeImportPermission(TransactionCase):

    def _user(self, login, groups):
        return self.env['res.users'].create({
            'name': login, 'login': login,
            'group_ids': [(6, 0, [self.env.ref(g).id for g in groups])]})

    def test_hr_manager_can_import(self):
        u = self._user('imp_mgr@test.vn', ['hr.group_hr_manager'])
        self.assertTrue(_cap_import_emp(self.env(user=u)))

    def test_admin_can_import(self):
        u = self._user('imp_admin@test.vn', ['base.group_system'])
        self.assertTrue(_cap_import_emp(self.env(user=u)))

    def test_hr_officer_cannot_import(self):
        u = self._user('imp_hr@test.vn', ['hr.group_hr_user'])
        self.assertFalse(_cap_import_emp(self.env(user=u)))

    def test_giaovu_cannot_import(self):
        u = self._user('imp_gv@test.vn', ['hocba_employees.group_hocba_giaovu'])
        self.assertFalse(_cap_import_emp(self.env(user=u)))


@tagged('post_install', '-at_install')
class TestEmployeeImportPreview(TransactionCase):

    def setUp(self):
        super().setUp()
        self.ctrl = HocBaEmployeeImport()
        http._request_stack.push(_FakeRequest(self.env))
        self.addCleanup(http._request_stack.pop)
        self.dep = self.env['hr.department'].create({'name': 'Phòng Test Import'})

    def test_catalogs_use_normalized_names(self):
        cat = self.ctrl._catalogs()
        self.assertEqual(cat['departments'].get('phong test import'), self.dep.id)

    def test_existing_collects_codes_and_cccds(self):
        emp = self.env['hr.employee'].create({
            'name': 'Đã có', 'x_employee_code': 'HB.XX'})
        emp.version_id.identification_id = '012345678901'
        found = self.ctrl._existing()
        self.assertIn('HB.XX', found['codes'])
        self.assertIn('012345678901', found['cccds'])

    def test_preview_creates_nothing(self):
        content = _book([FULL_HDR, _row(dep='Phòng Test Import')])
        before = self.env['hr.employee'].search_count([])
        res = self.ctrl._preview_payload(_FakeUpload(content), None)
        self.assertEqual(res['summary']['ok'], 1)
        self.assertEqual(self.env['hr.employee'].search_count([]), before)

    def test_preview_reports_existing_row_as_skipped(self):
        emp = self.env['hr.employee'].create({
            'name': 'Trùng', 'x_employee_code': 'HB.DUP'})
        content = _book([FULL_HDR, _row(code='HB.DUP', dep='Phòng Test Import')])
        res = self.ctrl._preview_payload(_FakeUpload(content), None)
        self.assertEqual(res['rows'], [])
        self.assertEqual(res['skipped'][0]['reason'], 'code_exists')
        self.assertTrue(emp.exists())

    def test_preview_rejects_non_xlsx(self):
        with self.assertRaises(EmployeeImportError) as cm:
            self.ctrl._preview_payload(_FakeUpload(b'x', filename='ns.csv'), None)
        self.assertEqual(cm.exception.code, 'bad_ext')

    def test_preview_rejects_missing_file(self):
        with self.assertRaises(EmployeeImportError) as cm:
            self.ctrl._preview_payload(None, None)
        self.assertEqual(cm.exception.code, 'no_file')

    def test_preview_keeps_filename(self):
        content = _book([FULL_HDR, _row(dep='Phòng Test Import')])
        res = self.ctrl._preview_payload(_FakeUpload(content, 'ds-nhan-su.xlsx'), None)
        self.assertEqual(res['filename'], 'ds-nhan-su.xlsx')


@tagged('post_install', '-at_install')
class TestEmployeeImportCommit(TransactionCase):

    def setUp(self):
        super().setUp()
        self.ctrl = HocBaEmployeeImport()
        http._request_stack.push(_FakeRequest(self.env))
        self.addCleanup(http._request_stack.pop)
        self.dep = self.env['hr.department'].create({'name': 'Phòng Test Import'})

    def _preview(self, data_rows):
        return self.ctrl._preview_payload(
            _FakeUpload(_book([FULL_HDR] + data_rows)), None)

    def test_commit_creates_employees_with_cccd_on_version(self):
        prev = self._preview([_row(code='HB.C1', dep='Phòng Test Import')])
        res = self.ctrl._commit_rows(prev['rows'])
        self.assertEqual(res['created'], 1)
        emp = self.env['hr.employee'].browse(res['employeeIds'][0])
        self.assertEqual(emp.x_employee_code, 'HB.C1')
        self.assertEqual(emp.department_id, self.dep)
        self.assertEqual(emp.version_id.identification_id, '038098029187')

    def test_commit_allows_official_missing_pit(self):
        prev = self._preview([_row(code='HB.C2', dep='Phòng Test Import')])
        res = self.ctrl._commit_rows(prev['rows'])
        emp = self.env['hr.employee'].browse(res['employeeIds'][0])
        self.assertEqual(emp.x_employment_status, 'official')
        self.assertEqual(res['needCompletion'], 1)

    def test_commit_rechecks_duplicate_even_if_client_lies(self):
        self.env['hr.employee'].create({'name': 'Có rồi', 'x_employee_code': 'HB.C3'})
        rows = [{'excelRow': 2, 'name': 'Nguyễn Văn A',
                 'values': {'name': 'Nguyễn Văn A', 'x_employee_code': 'HB.C3'},
                 'missingOfficial': []}]
        with self.assertRaises(EmployeeImportError) as cm:
            self.ctrl._commit_rows(rows)
        self.assertEqual(cm.exception.code, 'code_exists')

    def test_commit_ignores_field_outside_whitelist(self):
        rows = [{'excelRow': 2, 'name': 'X',
                 'values': {'name': 'X', 'wage': 99999999},
                 'missingOfficial': []}]
        res = self.ctrl._commit_rows(rows)
        emp = self.env['hr.employee'].browse(res['employeeIds'][0])
        self.assertNotEqual(emp.version_id.wage, 99999999)

    def test_commit_rejects_empty_name(self):
        rows = [{'excelRow': 5, 'name': '', 'values': {'name': '  '},
                 'missingOfficial': []}]
        with self.assertRaises(EmployeeImportError) as cm:
            self.ctrl._commit_rows(rows)
        self.assertEqual(cm.exception.code, 'empty_name')

    def test_probation_row_gets_no_onboarding_steps(self):
        prev = self._preview([_row(code='HB.C4', status='Thử việc',
                                   dep='Phòng Test Import', prob='01/06/2024')])
        res = self.ctrl._commit_rows(prev['rows'])
        emp = self.env['hr.employee'].browse(res['employeeIds'][0])
        self.assertEqual(emp.x_employment_status, 'probation')
        self.assertFalse(emp.x_onboarding_step_ids)

    def test_commit_is_all_or_nothing(self):
        good = {'excelRow': 2, 'name': 'Tốt',
                'values': {'name': 'Tốt', 'x_employee_code': 'HB.C5'},
                'missingOfficial': []}
        bad = {'excelRow': 3, 'name': '', 'values': {'name': ''},
               'missingOfficial': []}
        before = self.env['hr.employee'].search_count([])
        # savepoint = đúng ngữ nghĩa cr.rollback() mà route dùng; Odoo cấm gọi
        # thẳng rollback trong test (hỏng cursor của chính test).
        with self.assertRaises(EmployeeImportError):
            with self.env.cr.savepoint():
                self.ctrl._commit_rows([good, bad])
        self.assertEqual(self.env['hr.employee'].search_count([]), before)

    def test_two_rows_sharing_a_code_rejected_at_commit(self):
        rows = [{'excelRow': 2, 'name': 'A',
                 'values': {'name': 'A', 'x_employee_code': 'HB.C6'},
                 'missingOfficial': []},
                {'excelRow': 3, 'name': 'B',
                 'values': {'name': 'B', 'x_employee_code': 'HB.C6'},
                 'missingOfficial': []}]
        with self.assertRaises(EmployeeImportError) as cm:
            self.ctrl._commit_rows(rows)
        self.assertEqual(cm.exception.code, 'code_exists')
