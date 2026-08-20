# ============================================================
# Nhập hồ sơ nhân viên từ Excel — parser thuần hàm.
# Owner: Việt. Spec: docs/superpowers/specs/2026-08-20-employee-excel-import-design.md
#
# Tách khỏi controller để test thẳng bằng hàm: dựng worksheet → dò header →
# chuẩn hoá từng ô → kiểm từng nhánh lỗi, không cần dựng HTTP request.
#
# Nguyên tắc: danh mục lạ (Tình trạng, Phòng ban) là LỖI DÒNG, không âm thầm
# rơi về mặc định — file thật có 52 dòng ghi "Online" ở cột Tình trạng, để rơi
# về 'probation' là 52 người bị gán nhầm quy trình nhận việc.
# ============================================================
import re
import unicodedata

# 10 MB chứ không phải 2 MB như luồng nhập lịch làm việc: file nhân sự thật của
# trung tâm (59 sheet, ảnh nhúng) đã 2,21 MB — trần 2 MB chặn đúng cái file mà
# tính năng này sinh ra để nhập.
MAX_XLSX_BYTES = 10 * 1024 * 1024
MAX_HEADER_SCAN = 10        # chỉ dò header trong ngần này dòng đầu
MIN_HEADER_HITS = 3         # dòng phải khớp ngần này tiêu đề mới coi là header


class EmployeeImportError(Exception):
    """Lỗi cả file — controller đổi thẳng ra JSON 400."""

    def __init__(self, code, message):
        super().__init__(message)
        self.code = code
        self.message = message


def strip_accents(txt):
    txt = txt.replace('đ', 'd').replace('Đ', 'D')
    return ''.join(c for c in unicodedata.normalize('NFD', txt)
                   if unicodedata.category(c) != 'Mn')


def norm_header(val):
    """Tên cột → khoá so sánh: bỏ dấu, hạ hoa-thường, gộp khoảng trắng, bỏ
    hậu tố "_text" mà file nghiệp vụ hay gắn ("Họ tên nhân sự_text")."""
    if val is None:
        return ''
    txt = strip_accents(str(val)).lower().replace('_', ' ')
    txt = re.sub(r'\s+', ' ', txt).strip()
    return re.sub(r'\s*text$', '', txt).strip()


# Bí danh cột → khoá nội bộ. Khoá 'emailAlt' là email cá nhân: chỉ dùng khi
# file KHÔNG có cột email công ty.
HEADER_ALIASES = {
    'ma nhan su': 'code', 'ma nhan vien': 'code', 'ma nv': 'code',
    'ho va ten': 'name', 'ho ten': 'name', 'ho ten nhan su': 'name',
    'ho va ten nhan vien': 'name',
    'tinh trang': 'status', 'trang thai': 'status',
    'hinh thuc': 'workForm', 'hinh thuc lam viec': 'workForm',
    'loai vi tri': 'posType',
    'phong ban': 'dep', 'phong': 'dep',
    'chuc danh': 'job', 'vi tri': 'job',
    'ngay thu viec': 'probStart',
    'ngay thang nam sinh': 'bday', 'ngay sinh': 'bday',
    'so can cuoc cong dan': 'cccd', 'cccd': 'cccd', 'so cccd': 'cccd',
    'ngay cap': 'idIssue', 'noi cap': 'idPlace',
    'so dien thoai': 'phone', 'sdt': 'phone',
    'email cong ty': 'email', 'email': 'email', 'email ca nhan': 'emailAlt',
    'so tai khoan': 'bankAccountNo', 'ngan hang': 'bankCode',
    'ma so thue tncn': 'pit', 'mst': 'pit',
    'so so bhxh': 'si',
    'so the bhyt': 'hi', 'noi dang ky bhyt': 'hiPlace',
    # "thường chú" là lỗi chính tả có thật trong file nghiệp vụ của trung tâm.
    'dia chi thuong tru': 'permStreet', 'dia chi thuong chu': 'permStreet',
    'dia chi hien tai': 'currStreet',
}


def _map_row(cells):
    """Một dòng ô → (colmap, unknown_cols). Cột trùng khoá thì cột TRÁI thắng."""
    colmap, unknown = {}, []
    for idx, val in enumerate(cells):
        if val is None or not str(val).strip():
            continue
        key = HEADER_ALIASES.get(norm_header(val))
        if not key:
            unknown.append(str(val).strip())
        elif key not in colmap:
            colmap[key] = idx
    return colmap, unknown


def locate_header(ws):
    """Tìm dòng tiêu đề trong MAX_HEADER_SCAN dòng đầu.

    Trả (số dòng 1-based, colmap, danh sách cột không nhận diện được).
    """
    for row_idx, cells in enumerate(
            ws.iter_rows(min_row=1, max_row=MAX_HEADER_SCAN, values_only=True), 1):
        colmap, unknown = _map_row(cells)
        if len(colmap) < MIN_HEADER_HITS:
            continue
        # Email cá nhân chỉ được dùng khi không có cột email công ty.
        if 'email' not in colmap and 'emailAlt' in colmap:
            colmap['email'] = colmap['emailAlt']
        colmap.pop('emailAlt', None)
        if 'name' not in colmap:
            raise EmployeeImportError(
                'no_name_col',
                'File không có cột "Họ và tên" — không xác định được nhân viên. '
                'Đổi tên cột thành "Họ và tên" rồi tải lên lại.')
        return row_idx, colmap, unknown
    raise EmployeeImportError(
        'no_header',
        'Không tìm thấy dòng tiêu đề trong %d dòng đầu. File cần một dòng ghi '
        'tên cột (Mã nhân sự, Họ và tên, Tình trạng, Phòng ban…).' % MAX_HEADER_SCAN)


# ---------------------------------------------------------------- giá trị
import io
from datetime import date, datetime

import openpyxl

# Nhãn trong file nghiệp vụ → mã Selection của model. Khoá đã qua norm_header.
STATUS_MAP = {
    'chinh thuc': 'official', 'thu viec': 'probation', 'tts': 'intern',
    'parttime': 'parttime', 'part time': 'parttime', 'part-time': 'parttime',
    'ctv': 'ctv', 'co van': 'advisor', 'nghi viec': 'resigned',
    'dang offboarding': 'exiting',
}
WORK_FORM_MAP = {'offline': 'offline', 'online': 'online'}
POS_TYPE_MAP = {'quan ly': 'manager', 'nhan vien': 'staff', 'ctv': 'ctv',
                'freelancer': 'freelancer', 'co van': 'advisor'}

# Bí danh phòng ban: tên trong file → tên trong danh mục (đã norm_header).
# "Phòng Nhân sự" CỐ Ý không có ở đây — chưa hỏi được Học Bá nó thuộc đơn vị
# nào, nên để rơi vào bad_department thay vì đoán (xem spec §7.3).
DEPT_ALIASES = {
    'phong r&d sp': 'san pham (r&d_sp)',
    'phong r&d_sp': 'san pham (r&d_sp)',
    'ke toan': 'ke toan_hcns',
}

DATE_FORMATS = ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y')
EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')

# Khoá parser → field hr.employee. 'cccd' xử lý riêng (nằm trên hr.version).
TEXT_FIELDS = {
    'code': 'x_employee_code', 'phone': 'work_phone',
    'idPlace': 'x_id_place_issue', 'bankAccountNo': 'x_bank_account_no',
    'bankCode': 'x_bank_code', 'pit': 'x_pit_code',
    'si': 'x_social_insurance_no', 'hi': 'x_health_insurance_no',
    'hiPlace': 'x_health_care_place', 'permStreet': 'x_permanent_street',
    'currStreet': 'x_current_street',
}
DATE_FIELDS = {'probStart': 'x_probation_start', 'bday': 'birthday',
               'idIssue': 'x_id_date_issue'}


def _text(val):
    if val is None:
        return ''
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def to_date(val):
    """Ô Excel → date. Nhận ô kiểu ngày lẫn chuỗi dd/mm/yyyy | yyyy-mm-dd."""
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    txt = _text(val)
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(txt, fmt).date()
        except ValueError:
            continue
    return None


def clean_cccd(val):
    """Chuỗi CCCD → 12 chữ số, hoặc None. Giữ số 0 đầu (ô Excel là text)."""
    digits = re.sub(r'\D', '', _text(val))
    return digits if len(digits) == 12 else None


def _parse_row(cells, colmap, excel_row, catalogs):
    """Một dòng dữ liệu → (row dict, error dict). Đúng một trong hai khác None."""
    def cell(key):
        idx = colmap.get(key)
        return cells[idx] if idx is not None and idx < len(cells) else None

    def err(code, message):
        return None, {'excelRow': excel_row, 'code': code, 'message': message}

    name = _text(cell('name'))
    if not name:
        return err('empty_name', 'Dòng %d: thiếu Họ và tên.' % excel_row)

    vals, warnings = {'name': name}, []

    raw_status = _text(cell('status'))
    if raw_status:
        status = STATUS_MAP.get(norm_header(raw_status))
        if not status:
            return err('bad_status',
                       'Dòng %d: Tình trạng "%s" không thuộc danh mục '
                       '(Chính thức / Thử việc / TTS / Part-time / CTV / '
                       'Cố vấn / Nghỉ việc).' % (excel_row, raw_status))
        vals['x_employment_status'] = status

    raw_dep = _text(cell('dep'))
    if raw_dep:
        key = norm_header(raw_dep)
        # Đích của bí danh cũng phải qua norm_header: khoá danh mục do
        # _catalogs() dựng đã chuẩn hoá, "Sản phẩm (R&D_SP)" ở đó là
        # "san pham (r&d sp)" chứ không còn gạch dưới.
        key = norm_header(DEPT_ALIASES.get(key, key))
        dep_id = (catalogs.get('departments') or {}).get(key)
        if not dep_id:
            return err('bad_department',
                       'Dòng %d: Phòng ban "%s" không khớp danh mục. Sửa lại '
                       'trong file hoặc thêm phòng ban vào hệ thống rồi tải '
                       'lên lại.' % (excel_row, raw_dep))
        vals['department_id'] = dep_id

    raw_job = _text(cell('job'))
    if raw_job:
        job_id = (catalogs.get('jobs') or {}).get(norm_header(raw_job))
        if job_id:
            vals['job_id'] = job_id
        # Luôn giữ nguyên văn: danh mục hệ thống ~16 vị trí, file có 48.
        vals['job_title'] = raw_job

    for key, code in (('workForm', 'x_work_form'), ('posType', 'x_position_type')):
        raw = _text(cell(key))
        if not raw:
            continue
        table = WORK_FORM_MAP if key == 'workForm' else POS_TYPE_MAP
        mapped = table.get(norm_header(raw))
        if mapped:
            vals[code] = mapped
        else:
            warnings.append('%s "%s" không thuộc danh mục — để trống.'
                            % ('Hình thức' if key == 'workForm' else 'Loại vị trí', raw))

    for key, field in DATE_FIELDS.items():
        raw = cell(key)
        if raw in (None, ''):
            continue
        parsed = to_date(raw)
        if parsed:
            vals[field] = parsed
        else:
            warnings.append('Không đọc được ngày "%s" — để trống.' % _text(raw))

    for key, field in TEXT_FIELDS.items():
        raw = _text(cell(key))
        if raw:
            vals[field] = raw

    raw_email = _text(cell('email'))
    if raw_email:
        if EMAIL_RE.match(raw_email):
            vals['work_email'] = raw_email
        else:
            warnings.append('Email "%s" sai định dạng — để trống.' % raw_email)

    raw_cccd = cell('cccd')
    if _text(raw_cccd):
        cccd = clean_cccd(raw_cccd)
        if cccd:
            vals['cccd'] = cccd
        else:
            warnings.append('CCCD "%s" không đủ 12 chữ số — để trống.'
                            % _text(raw_cccd))

    missing = []
    if vals.get('x_employment_status') == 'official':
        if not vals.get('cccd'):
            missing.append('CCCD')
        if not vals.get('x_pit_code'):
            missing.append('MST TNCN')
        if not vals.get('x_social_insurance_no'):
            missing.append('Số sổ BHXH')

    return {
        'excelRow': excel_row,
        'name': name,
        'code': vals.get('x_employee_code') or '',
        'depName': raw_dep,
        'jobName': raw_job,
        'status': vals.get('x_employment_status') or '',
        'warnings': warnings,
        'missingOfficial': missing,
        'values': vals,
    }, None


def parse_employees_xlsx(content, sheet=None, catalogs=None, existing=None):
    """Đọc + KIỂM file, KHÔNG ghi gì. Trả bảng xem trước cho SPA."""
    if len(content) > MAX_XLSX_BYTES:
        raise EmployeeImportError(
            'too_large', 'File nặng quá %d MB.' % (MAX_XLSX_BYTES // (1024 * 1024)))
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise EmployeeImportError(
            'bad_file', 'Không mở được file. Hãy mở bằng Excel rồi "Lưu dưới '
                        'dạng" .xlsx và tải lên lại.')

    catalogs = catalogs or {}
    existing = existing or {}
    codes = set(existing.get('codes') or ())
    cccds = set(existing.get('cccds') or ())

    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.worksheets[0]
    header_row, colmap, unknown = locate_header(ws)

    rows, errors, skipped = [], [], []
    for offset, cells in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True)):
        if not any(_text(c) for c in cells):
            continue
        excel_row = header_row + 1 + offset
        row, error = _parse_row(cells, colmap, excel_row, catalogs)
        if error:
            errors.append(error)
            continue
        code, cccd = row['values'].get('x_employee_code'), row['values'].get('cccd')
        if code and code in codes:
            skipped.append({'excelRow': excel_row, 'code': code,
                            'name': row['name'], 'reason': 'code_exists'})
            continue
        if cccd and cccd in cccds:
            skipped.append({'excelRow': excel_row, 'code': code or '',
                            'name': row['name'], 'reason': 'cccd_exists'})
            continue
        if code:
            codes.add(code)
        if cccd:
            cccds.add(cccd)
        rows.append(row)

    return {
        'sheets': list(wb.sheetnames),
        'sheet': ws.title,
        'headerRow': header_row,
        'rows': rows,
        'skipped': skipped,
        'errors': errors,
        'unknownCols': unknown,
        'summary': {
            'total': len(rows) + len(skipped) + len(errors),
            'ok': len(rows),
            'skipped': len(skipped),
            'error': len(errors),
            'needCompletion': sum(1 for r in rows if r['missingOfficial']),
        },
    }
