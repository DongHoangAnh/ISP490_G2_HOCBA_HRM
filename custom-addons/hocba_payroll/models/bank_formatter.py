"""
Bank file formatters — Strategy Pattern.
FUNC-PR-003: Mỗi ngân hàng là 1 concrete strategy.

    Usage:
        formatter = BankFormatterRegistry.get('VCB')
        rows = formatter.build_rows(payslips, description_tpl)
        file_bytes = formatter.export_xlsx(rows)
"""
import io
import logging
import re
import unicodedata

from odoo import _
from odoo.exceptions import ValidationError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, numbers
except ImportError:
    openpyxl = None
    _logger.warning('openpyxl not installed — bank file export will fail.')


# ── Helpers ──────────────────────────────────────────────────
def _remove_diacritics(text):
    """Remove Vietnamese diacritics: Nguyễn → Nguyen."""
    if not text:
        return ''
    nfkd = unicodedata.normalize('NFKD', text)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def _to_uppercase_no_diacritics(text):
    """NGUYEN VAN A — required by VCB."""
    return _remove_diacritics(text).upper()


# ═════════════════════════════════════════════════════════════
# Abstract base (Strategy interface)
# ═════════════════════════════════════════════════════════════
class BaseBankFormatter:
    """Base class for all bank formatters."""

    code = ''           # e.g. 'VCB'
    name = ''           # e.g. 'Vietcombank'
    headers = []        # list[str] — column headers

    def validate_account(self, acc_number, regex=None):
        """Validate account number format."""
        if not acc_number:
            return False
        if regex and not re.match(regex, acc_number.strip()):
            return False
        return True

    def build_row(self, stt, acc_number, emp_name, amount, description, **kw):
        """Build a single data row — override per bank."""
        raise NotImplementedError

    def build_rows(self, payslips, description_tpl, account_regex=None):
        """Build all rows from a set of payslips.

        Returns:
            (list[dict], list[str])  — rows, warnings
        """
        rows = []
        warnings = []
        stt = 1
        for slip in payslips:
            net = self._get_net_amount(slip)
            if net <= 0:
                warnings.append(
                    _('Payslip %(ref)s có Net = 0 hoặc âm — đã bỏ qua.',
                      ref=slip.number or slip.name)
                )
                continue

            bank_account = self._get_employee_bank(slip.employee_id)
            if not bank_account or not bank_account.acc_number:
                continue  # already validated in wizard

            acc_number = bank_account.acc_number.strip()
            emp_name = slip.employee_id.name or ''
            emp_code = getattr(slip.employee_id, 'x_employee_code', '') or ''
            month = slip.date_to.strftime('%m') if slip.date_to else ''
            year = slip.date_to.strftime('%Y') if slip.date_to else ''

            desc = (description_tpl or 'Luong T{month}/{year}').format(
                month=month, year=year, employee_code=emp_code,
            )

            row = self.build_row(
                stt=stt,
                acc_number=acc_number,
                emp_name=emp_name,
                amount=int(net),
                description=desc,
                employee_code=emp_code,
            )
            rows.append(row)
            stt += 1

        return rows, warnings

    def export_xlsx(self, rows):
        """Generate XLSX bytes from rows."""
        if openpyxl is None:
            raise ValidationError(_('Thư viện openpyxl chưa được cài đặt.'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.code

        # Header row
        header_font = Font(bold=True)
        for col_idx, header in enumerate(self.headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font

        # Data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, key in enumerate(self._column_keys(), start=1):
                val = row_data.get(key, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _column_keys(self):
        """Return ordered keys matching self.headers."""
        raise NotImplementedError

    @staticmethod
    def _get_net_amount(payslip):
        """Read NET amount from payslip lines."""
        net_line = payslip.line_ids.filtered(lambda l: l.code == 'thuc_lanh')
        return net_line[0].amount if net_line else 0.0

    @staticmethod
    def _get_employee_bank(employee):
        """Get employee bank account — compatible with Community (no bank_account_id)."""
        if hasattr(employee, 'bank_account_id') and employee.bank_account_id:
            return employee.bank_account_id
        # Community fallback: partner bank accounts
        partner = (
            getattr(employee, 'address_home_id', None)
            or getattr(employee, 'work_contact_id', None)
        )
        if partner and partner.bank_ids:
            return partner.bank_ids[0]
        return None


# ═════════════════════════════════════════════════════════════
# Concrete: Vietcombank (VCB)
# ═════════════════════════════════════════════════════════════
class VCBFormatter(BaseBankFormatter):
    code = 'VCB'
    name = 'Vietcombank'
    headers = ['STT', 'Số tài khoản', 'Tên người nhận', 'Số tiền', 'Nội dung']

    def build_row(self, stt, acc_number, emp_name, amount, description, **kw):
        return {
            'stt': stt,
            'acc_number': acc_number,
            'emp_name': _to_uppercase_no_diacritics(emp_name),   # BR-PR-015
            'amount': amount,
            'description': description,
        }

    def _column_keys(self):
        return ['stt', 'acc_number', 'emp_name', 'amount', 'description']


# ═════════════════════════════════════════════════════════════
# Concrete: Techcombank (TCB)
# ═════════════════════════════════════════════════════════════
class TCBFormatter(BaseBankFormatter):
    code = 'TCB'
    name = 'Techcombank'
    headers = ['Account Number', 'Beneficiary Name', 'Amount', 'Currency', 'Remark', 'Bank Code']

    def build_row(self, stt, acc_number, emp_name, amount, description, **kw):
        return {
            'acc_number': acc_number,
            'emp_name': emp_name.upper(),    # BR-PR-016: giữ dấu TV, viết hoa
            'amount': amount,
            'currency': 'VND',
            'description': description[:100],  # VR-009: max 100 chars
            'bank_code': '',
        }

    def _column_keys(self):
        return ['acc_number', 'emp_name', 'amount', 'currency', 'description', 'bank_code']


# ═════════════════════════════════════════════════════════════
# Concrete: MB Bank (eMB_BulkPayment format)
# ═════════════════════════════════════════════════════════════
class MBBankFormatter(BaseBankFormatter):
    """MB Bank bulk payment format (eMB_BulkPayment).

    Columns:
        A — STT (Ord. No.)
        B — Số tài khoản (Account No.)
        C — Tên đơn vị thụ hưởng (Beneficiary Organization)
        D — Ngân hàng thụ hưởng/Chi nhánh (Beneficiary Bank)
        E — Số tiền (Amount)
        F — Chi tiết thanh toán (Payment Detail)
    """

    code = 'MB'
    name = 'MB Bank'
    headers = [
        '\ufeffSTT\n(Ord. No.)\n(1)',
        'Số tài khoản\n(Account No.)\n(2)',
        'Tên đơn vị thụ hưởng\n(Beneficiary Organization)\n(3)',
        'Ngân hàng thụ hưởng/Chi nhánh\n(Beneficiary Bank)\n(4)',
        'Số tiền\n(Amount)\n(5)',
        'Chi tiết thanh toán\n(Payment Detail)\n(6)',
    ]

    def _build_bank_lookup(self, env):
        """Build a lookup dict: short_code → full MB bank name."""
        entries = env['hb.bank.format'].sudo().search([('active', '=', True)])
        lookup = {}
        for entry in entries:
            # Extract short code from "CODE - Full name"
            parts = entry.name.split(' - ', 1)
            if len(parts) == 2:
                code = parts[0].strip().upper()
                if code not in lookup:
                    lookup[code] = entry.name
        return lookup

    def _resolve_bank_name(self, bank_account, lookup):
        """Resolve employee bank account to MB-format bank name."""
        if not bank_account:
            return ''
        bank = bank_account.bank_id
        if not bank:
            return ''

        bank_name = bank.name or ''
        bank_bic = (bank.bic or '').upper()

        # Try matching by BIC prefix (e.g., BFTV → BIDV)
        # Try matching by bank name against lookup codes
        for code, full_name in lookup.items():
            # Check if short code appears in BIC
            if bank_bic and code in bank_bic:
                return full_name
            # Check if short code matches bank name (case-insensitive)
            if code.lower() in bank_name.lower():
                return full_name
            # Check if bank name appears in the full MB entry name
            if bank_name.lower() in full_name.lower():
                return full_name

        # Fallback: return raw bank name
        return bank_name

    def build_row(self, stt, acc_number, emp_name, amount, description, **kw):
        bank_name = kw.get('bank_name', '')
        return {
            'stt': stt,
            'acc_number': acc_number,
            'emp_name': _remove_diacritics(emp_name),
            'bank_name': bank_name,
            'amount': amount,
            'description': _remove_diacritics(description),
        }

    def _column_keys(self):
        return ['stt', 'acc_number', 'emp_name', 'bank_name', 'amount', 'description']

    def build_rows(self, payslips, description_tpl, account_regex=None):
        """Override to include bank name lookup for column D."""
        rows = []
        warnings = []
        stt = 1

        # Build bank lookup from DB (via payslip env)
        lookup = {}
        if payslips:
            lookup = self._build_bank_lookup(payslips[0].env)

        for slip in payslips:
            net = self._get_net_amount(slip)
            if net <= 0:
                warnings.append(
                    _('Payslip %(ref)s có Net = 0 hoặc âm — đã bỏ qua.',
                      ref=slip.number or slip.name)
                )
                continue

            bank_account = self._get_employee_bank(slip.employee_id)
            if not bank_account or not bank_account.acc_number:
                continue

            acc_number = bank_account.acc_number.strip()
            emp_name = slip.employee_id.name or ''
            emp_code = getattr(slip.employee_id, 'x_employee_code', '') or ''
            month = slip.date_to.strftime('%m') if slip.date_to else ''
            year = slip.date_to.strftime('%Y') if slip.date_to else ''

            desc = (description_tpl or 'Luong T{month}/{year}').format(
                month=month, year=year, employee_code=emp_code,
            )

            bank_name = self._resolve_bank_name(bank_account, lookup)

            row = self.build_row(
                stt=stt,
                acc_number=acc_number,
                emp_name=emp_name,
                amount=int(net),
                description=desc,
                bank_name=bank_name,
                employee_code=emp_code,
            )
            rows.append(row)
            stt += 1

        return rows, warnings

    def export_xlsx(self, rows):
        """Generate eMB_BulkPayment format XLSX."""
        if openpyxl is None:
            raise ValidationError(_('Thư viện openpyxl chưa được cài đặt.'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'eMB_BulkPayment'

        # Row 1: Title in B1
        ws.merge_cells('B1:F1')
        title_cell = ws['B1']
        title_cell.value = 'DANH SÁCH GIAO DỊCH\n(LIST OF TRANSACTIONS)'
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True,
        )
        ws.row_dimensions[1].height = 40

        # Row 2: Column headers
        header_font = Font(bold=True, size=10)
        header_align = Alignment(
            horizontal='center', vertical='center', wrap_text=True,
        )
        for col_idx, header in enumerate(self.headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = header_align
        ws.row_dimensions[2].height = 50

        # Data rows starting from row 3
        for row_idx, row_data in enumerate(rows, start=3):
            for col_idx, key in enumerate(self._column_keys(), start=1):
                val = row_data.get(key, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')

        # Column widths matching eMB template
        ws.column_dimensions['A'].width = 8    # STT
        ws.column_dimensions['B'].width = 22   # Số tài khoản
        ws.column_dimensions['C'].width = 30   # Tên thụ hưởng
        ws.column_dimensions['D'].width = 55   # Ngân hàng thụ hưởng
        ws.column_dimensions['E'].width = 18   # Số tiền
        ws.column_dimensions['F'].width = 40   # Chi tiết thanh toán

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


# ═════════════════════════════════════════════════════════════
# Registry (simple factory)
# ═════════════════════════════════════════════════════════════
class BankFormatterRegistry:
    """Registry for bank formatter strategies."""

    _formatters = {
        'VCB': VCBFormatter,
        'TCB': TCBFormatter,
        'MB': MBBankFormatter,
    }

    @classmethod
    def get(cls, code):
        """Return an instance of the formatter for the given bank code."""
        formatter_cls = cls._formatters.get(code)
        if not formatter_cls:
            raise ValidationError(
                _('Không tìm thấy formatter cho ngân hàng "%(code)s".', code=code)
            )
        return formatter_cls()

    @classmethod
    def register(cls, code, formatter_cls):
        """Register a new formatter (for extensibility)."""
        cls._formatters[code] = formatter_cls

    @classmethod
    def available_codes(cls):
        return list(cls._formatters.keys())
