import odoo
import odoo.tools
import odoo.api
from odoo.modules.registry import Registry
from datetime import datetime, date, timedelta
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def seed_july_attendance():
    odoo.tools.config.parse_config(['-c', '/etc/odoo/odoo.conf', '-d', 'neondb'])
    reg = Registry('neondb')
    
    with reg.cursor() as cr:
        env = odoo.api.Environment(cr, odoo.SUPERUSER_ID, {})
        
        # Get active employees
        employees = env['hr.employee'].search([('active', '=', True)])
        print(f"Total active employees: {len(employees)}")

        # Clear existing July 2026 attendance if any
        existing = env['hocba.attendance'].search([
            ('check_in', '>=', '2026-07-01 00:00:00'),
            ('check_in', '<=', '2026-07-31 23:59:59')
        ])
        if existing:
            print(f"Deleting {len(existing)} existing July attendance records...")
            existing.unlink()

        # July 2026 working days (Mon-Fri)
        july_days = []
        for day in range(1, 32):
            d = date(2026, 7, day)
            if d.weekday() < 5:  # Monday=0, ..., Friday=4
                july_days.append(d)

        print(f"July 2026 has {len(july_days)} working days (Mon-Fri).")

        records_to_create = []
        excel_rows = []
        stt = 1

        for emp in employees:
            # Deterministic seed per employee so data is consistent
            random.seed(emp.id + 202607)
            
            for d in july_days:
                # 95% full day attendance, 5% late or half-day
                rand_val = random.random()
                
                if rand_val > 0.05:
                    # Full day standard (08:00 - 17:00 ICT = 01:00 - 10:00 UTC)
                    check_in_dt = datetime(d.year, d.month, d.day, 1, 0, 0)
                    check_out_dt = datetime(d.year, d.month, d.day, 10, 0, 0)
                    work_credit = 1.0
                    working_hours = 8.0
                    late_mins = 0
                    early_leave_mins = 0
                    note = "Đúng giờ"
                elif rand_val > 0.02:
                    # 10-15 mins late
                    check_in_dt = datetime(d.year, d.month, d.day, 1, 15, 0)
                    check_out_dt = datetime(d.year, d.month, d.day, 10, 0, 0)
                    work_credit = 1.0
                    working_hours = 7.75
                    late_mins = 15
                    early_leave_mins = 0
                    note = "Đi trễ 15p"
                else:
                    # Half day morning
                    check_in_dt = datetime(d.year, d.month, d.day, 1, 0, 0)
                    check_out_dt = datetime(d.year, d.month, d.day, 5, 0, 0)
                    work_credit = 0.5
                    working_hours = 4.0
                    late_mins = 0
                    early_leave_mins = 0
                    note = "Làm nửa ngày (Sáng)"

                records_to_create.append({
                    'employee_id': emp.id,
                    'check_in': check_in_dt.strftime('%Y-%m-%d %H:%M:%S'),
                    'check_out': check_out_dt.strftime('%Y-%m-%d %H:%M:%S'),
                    'work_credit': work_credit,
                    'working_hours': working_hours,
                    'morning_credit': 0.5 if work_credit >= 0.5 else 0.0,
                    'afternoon_credit': 0.5 if work_credit == 1.0 else 0.0,
                    'late_minutes': late_mins,
                    'early_leave_minutes': early_leave_mins,
                    'notes': note,
                })

                excel_rows.append((
                    stt,
                    emp.barcode or f"NV{emp.id:04d}",
                    emp.name,
                    emp.department_id.name or "Phòng ban chung",
                    d.strftime('%d/%m/%Y'),
                    (check_in_dt + timedelta(hours=7)).strftime('%H:%M'),
                    (check_out_dt + timedelta(hours=7)).strftime('%H:%M'),
                    work_credit,
                    working_hours,
                    late_mins,
                    note
                ))
                stt += 1

        # Bulk create in chunks of 500
        total_records = len(records_to_create)
        print(f"Creating {total_records} attendance records in database...")
        
        chunk_size = 500
        for i in range(0, total_records, chunk_size):
            chunk = records_to_create[i:i+chunk_size]
            env['hocba.attendance'].create(chunk)
            print(f"Created {i + len(chunk)} / {total_records} records...")

        cr.commit()
        print("SUCCESS: Database attendance records for July 2026 created!")

        # Export Excel file
        excel_path = "/mnt/extra-addons/Mau_Cham_Cong_Thang_07_2026.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chấm công T07-2026"
        ws.views.sheetView[0].showGridLines = True

        ws.merge_cells("A1:K1")
        title_cell = ws["A1"]
        title_cell.value = "BẢNG DỮ LIỆU CHẤM CÔNG NHÂN VIÊN — THÁNG 07/2026"
        title_cell.font = Font(name="Arial", size=14, bold=True, color="003366")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        ws.merge_cells("A2:K2")
        sub_cell = ws["A2"]
        sub_cell.value = f"Công ty: Học Bá Education | Số lượng: {len(employees)} nhân viên | Tổng bản ghi: {total_records}"
        sub_cell.font = Font(name="Arial", size=10, italic=True, color="555555")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        headers = [
            "STT", "Mã NV", "Họ và Tên", "Phòng ban", "Ngày", 
            "Check In", "Check Out", "Số công", "Số giờ làm", "Đi trễ (Phút)", "Ghi chú"
        ]
        
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[4].height = 25
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        thin_border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )

        for row_idx, row_data in enumerate(excel_rows, 5):
            ws.row_dimensions[row_idx].height = 20
            bg_color = "FFFFFF" if row_idx % 2 == 1 else "F9FAFB"
            row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
            for col_idx, val in enumerate(row_data, 1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.fill = row_fill
                c.border = thin_border
                c.font = Font(name="Arial", size=10)
                if col_idx in (1, 2, 5, 6, 7, 8, 9, 10):
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")

        wb.save(excel_path)
        print(f"Saved Excel template to {excel_path}")

if __name__ == "__main__":
    seed_july_attendance()
