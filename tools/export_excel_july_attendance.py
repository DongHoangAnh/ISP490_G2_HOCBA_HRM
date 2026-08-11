import odoo
import odoo.tools
import odoo.api
from odoo.modules.registry import Registry
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def export_july_excel():
    odoo.tools.config.parse_config(['-c', '/etc/odoo/odoo.conf', '-d', 'neondb'])
    reg = Registry('neondb')
    
    with reg.cursor() as cr:
        env = odoo.api.Environment(cr, odoo.SUPERUSER_ID, {})
        
        atts = env['hocba.attendance'].search([
            ('check_in', '>=', '2026-07-01 00:00:00'),
            ('check_in', '<=', '2026-07-31 23:59:59')
        ], order='date asc, employee_id asc')
        
        print(f"Exporting {len(atts)} attendance records to Excel...")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chấm công T07-2026"
        ws.views.sheetView[0].showGridLines = True

        # Header Title
        ws.merge_cells("A1:K1")
        title_cell = ws["A1"]
        title_cell.value = "BẢNG DỮ LIỆU CHẤM CÔNG NHÂN VIÊN — THÁNG 07/2026"
        title_cell.font = Font(name="Arial", size=14, bold=True, color="003366")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        ws.merge_cells("A2:K2")
        sub_cell = ws["A2"]
        sub_cell.value = f"Công ty: Học Bá Education | Tổng bản ghi chấm công: {len(atts)}"
        sub_cell.font = Font(name="Arial", size=10, italic=True, color="555555")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        headers = [
            "STT", "Mã NV", "Họ và Tên", "Phòng ban", "Ngày", 
            "Check In", "Check Out", "Số công", "Số giờ làm", "Đi trễ (Phút)", "Ghi chú"
        ]
        
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[4].height = 25
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        thin_border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )

        for idx, a in enumerate(atts, 1):
            row_idx = idx + 4
            ws.row_dimensions[row_idx].height = 20
            bg_color = "FFFFFF" if idx % 2 == 1 else "F9FAFB"
            row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
            in_ict = (a.check_in + timedelta(hours=7)).strftime('%H:%M') if a.check_in else ""
            out_ict = (a.check_out + timedelta(hours=7)).strftime('%H:%M') if a.check_out else ""
            date_str = (a.check_in + timedelta(hours=7)).strftime('%d/%m/%Y') if a.check_in else ""

            row_data = (
                idx,
                a.employee_id.barcode or f"NV{a.employee_id.id:04d}",
                a.employee_id.name,
                a.employee_id.department_id.name or "Phòng ban chung",
                date_str,
                in_ict,
                out_ict,
                a.work_credit,
                a.working_hours,
                a.late_minutes,
                a.notes or ""
            )

            for col_idx, val in enumerate(row_data, 1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.fill = row_fill
                c.border = thin_border
                c.font = Font(name="Arial", size=10)
                if col_idx in (1, 2, 5, 6, 7, 8, 9, 10):
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center")

        excel_path = "/mnt/extra-addons/Mau_Cham_Cong_Thang_07_2026.xlsx"
        wb.save(excel_path)
        print(f"Saved Excel file to {excel_path}")

if __name__ == "__main__":
    export_july_excel()
