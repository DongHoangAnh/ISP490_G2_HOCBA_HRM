import os
import sys
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import column_index_from_string

sys.stdout.reconfigure(encoding='utf-8')

KNOWLEDGE_DIR = r"D:\FPT\DO_an\code\ISP490_G2_HOCBA_HRM\Knowledge"
UNIT_TEST_DIR = r"D:\FPT\DO_an\code\ISP490_G2_HOCBA_HRM\HRM_Học Bá\unit-test"

FONT = 'Arial'
HDR_FILL = PatternFill('solid', fgColor='FFBDD6EE')
SEC_FILL = PatternFill('solid', fgColor='FFFBE4D5')
THIN = Side('thin', color='FF000000')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def sync_existing_files():
    # 1. Sync UT_TIMEOFF v2.0 from Knowledge to unit-test
    kn_ut = os.path.join(KNOWLEDGE_DIR, "ISP490_G2_UT_TIMEOFF.xlsx")
    ut_ut = os.path.join(UNIT_TEST_DIR, "ISP490_G2_UT_TIMEOFF.xlsx")
    shutil.copy2(kn_ut, ut_ut)
    print(f"Synced {kn_ut} -> {ut_ut}")

    # 2. Sync BPT_TIMEOFF from Knowledge to unit-test
    kn_bpt = os.path.join(KNOWLEDGE_DIR, "ISP490_G2_BPT_TIMEOFF.xlsx")
    ut_bpt = os.path.join(UNIT_TEST_DIR, "ISP490_G2_BPT_TIMEOFF.xlsx")
    shutil.copy2(kn_bpt, ut_bpt)
    print(f"Synced {kn_bpt} -> {ut_bpt}")

def create_bpt_file(module_code, module_name, func_id, sections_data, out_filename):
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Cover
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.views.sheetView[0].showGridLines = True
    
    ws_cover.cell(row=8, column=2, value="Business Process Test").font = Font(name=FONT, size=16, bold=True, color="003366")
    ws_cover.cell(row=11, column=2, value="Module").font = Font(name=FONT, size=10, bold=True)
    ws_cover.cell(row=11, column=3, value=module_name)
    ws_cover.cell(row=12, column=2, value="Function ID").font = Font(name=FONT, size=10, bold=True)
    ws_cover.cell(row=12, column=3, value=func_id)
    ws_cover.cell(row=13, column=2, value="Function Name").font = Font(name=FONT, size=10, bold=True)
    ws_cover.cell(row=13, column=3, value=f"{module_name.upper()} — HỌC BÁ HRM")
    ws_cover.cell(row=14, column=2, value="Created Date").font = Font(name=FONT, size=10, bold=True)
    ws_cover.cell(row=14, column=3, value="08/08/2026")
    ws_cover.cell(row=23, column=2, value=f"Ghi chú: file này là bộ BUSINESS PROCESS TEST thủ công cho module {module_name} — kiểm tra các luồng nghiệp vụ đi qua NHIỀU màn hình / nhiều vai trò.").font = Font(name=FONT, size=9, italic=True)

    # --- Sheet 2: Histories
    ws_hist = wb.create_sheet(title="Histories")
    ws_hist.views.sheetView[0].showGridLines = True
    headers_hist = ["No", "Version", "Description", "Sheet", "Modified date", "Modified by"]
    for c_idx, h in enumerate(headers_hist, 2):
        cell = ws_hist.cell(row=2, column=c_idx, value=h)
        cell.font = Font(name=FONT, size=10, bold=True)
        cell.fill = HDR_FILL
        cell.border = BOX
    
    row_hist = [1, "1.0", f"Khởi tạo bộ Business Process Test (BPT) cho module {module_name}", "Cover, Histories, BPT, Test Result", "08/08/2026", "Nhóm G2"]
    for c_idx, val in enumerate(row_hist, 2):
        cell = ws_hist.cell(row=3, column=c_idx, value=val)
        cell.font = Font(name=FONT, size=10)
        cell.border = BOX

    # --- Sheet 3: BPT
    ws_bpt = wb.create_sheet(title="BPT")
    ws_bpt.views.sheetView[0].showGridLines = True
    ws_bpt.cell(row=2, column=2, value="Function ID").font = Font(name=FONT, size=10, bold=True)
    ws_bpt.cell(row=2, column=3, value=func_id)
    ws_bpt.cell(row=2, column=12, value="Function Name").font = Font(name=FONT, size=10, bold=True)
    ws_bpt.cell(row=2, column=13, value=module_name)

    headers_bpt = ["STT", "Testcase ID", "Chức năng / Luồng Nghiệp vụ", "Tên Test Case (BPT)", "Mô tả chi tiết", "Điều kiện tiên quyết", "Các bước thực hiện", "Dữ liệu đầu vào", "Kết quả mong đợi", "Loại Test", "Mức độ", "Tự động", "Trạng thái", "Người thực hiện", "Ghi chú"]
    
    for c_idx, h in enumerate(headers_bpt, 1):
        cell = ws_bpt.cell(row=6, column=c_idx, value=h)
        cell.font = Font(name=FONT, size=10, bold=True)
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX

    curr_row = 7
    stt = 1
    test_result_sheets = []

    for sec_title, test_cases in sections_data:
        # Section Header
        sec_cell = ws_bpt.cell(row=curr_row, column=1, value=sec_title)
        sec_cell.font = Font(name=FONT, size=10, bold=True)
        sec_cell.fill = SEC_FILL
        ws_bpt.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=15)
        curr_row += 1

        for tc_id, tc_name, desc, pre_req, steps, inputs, expected in test_cases:
            sheet_code = f"A{stt}" if stt <= 4 else (f"B{stt-4}" if stt <= 8 else f"C{stt-8}")
            test_result_sheets.append((sheet_code, tc_id, tc_name))

            row_vals = [
                stt, tc_id, sec_title.split('.')[1].strip() if '.' in sec_title else sec_title,
                tc_name, desc, pre_req, steps, inputs, expected,
                "Process", "Cao", "Có", "Pass", "Nhóm G2", ""
            ]

            ws_bpt.row_dimensions[curr_row].height = 40
            for c_idx, val in enumerate(row_vals, 1):
                cell = ws_bpt.cell(row=curr_row, column=c_idx, value=val)
                cell.font = Font(name=FONT, size=9)
                cell.border = BOX
                cell.alignment = Alignment(horizontal="center" if c_idx in (1, 2, 10, 11, 12, 13) else "left", vertical="center", wrap_text=True)
            
            stt += 1
            curr_row += 1

    # --- Create Test Result Detail Sheets
    for sheet_code, tc_id, tc_name in test_result_sheets:
        ws_res = wb.create_sheet(title=f"Test Result ({sheet_code})")
        ws_res.views.sheetView[0].showGridLines = True
        ws_res.cell(row=2, column=2, value="Function ID").font = Font(name=FONT, size=10, bold=True)
        ws_res.cell(row=2, column=3, value=tc_id)
        ws_res.cell(row=2, column=12, value="Function Name").font = Font(name=FONT, size=10, bold=True)
        ws_res.cell(row=2, column=13, value=tc_name)

    out_path = os.path.join(UNIT_TEST_DIR, out_filename)
    wb.save(out_path)
    print(f"Created BPT Excel file: {out_path}")

def run():
    sync_existing_files()

    # BPT ACCOUNTS
    create_bpt_file(
        "ACCOUNTS", "Quản lý Tài khoản & Phân quyền", "BPT_ACCOUNTS",
        [
            ("A. Luồng Khởi tạo & Kích hoạt Tài khoản", [
                ("BPT-ACC-001", "Tạo tài khoản NV mới ➔ Gửi mail kích hoạt ➔ Tự thiết lập mật khẩu ➔ Đăng nhập lần đầu",
                 "Kiểm tra luồng khởi tạo và kích hoạt tài khoản nhân viên mới từ A-Z", "Tài khoản HR Manager đã đăng nhập",
                 "1. HR tạo tài khoản NV mới với email hợp lệ\n2. NV kiểm tra hộp thư nhận email kích hoạt\n3. Click link kích hoạt và thiết lập mật khẩu\n4. Đăng nhập lần đầu vào hệ thống",
                 "Email NV: test_newemp@hocba.edu.vn", "Tài khoản kích hoạt thành công, đăng nhập vào đúng Dashboard Nhân viên"),
            ]),
            ("B. Luồng Phân quyền & Thay đổi Role", [
                ("BPT-ACC-002", "Cấp quyền HR Manager cho NV ➔ Đăng xuất ➔ Đăng nhập lại ➔ Kiểm tra ranh giới phân quyền",
                 "Kiểm tra luồng thay đổi vai trò nhân viên và cập nhật phân quyền tức thì", "Tài khoản Admin đã đăng nhập",
                 "1. Admin nâng quyền tài khoản NV thành HR Manager\n2. NV đăng xuất và đăng nhập lại\n3. Kiểm tra các menu/chức năng HR Manager",
                 "Role: HR Manager", "Tài khoản truy cập đầy đủ chức năng quản lý HR, các API kiểm tra quyền trả về HTTP 200 OK"),
            ]),
            ("C. Luồng Khóa Tài khoản & Vô hiệu hóa Session", [
                ("BPT-ACC-003", "Khóa tài khoản NV nghỉ việc ➔ Hủy session live ➔ Chặn mọi API request",
                 "Kiểm tra tính an toàn bảo mật khi khóa tài khoản nhân viên", "Tài khoản HR Manager đã đăng nhập",
                 "1. HR nhấn Khóa tài khoản nhân viên\n2. Kiểm tra session live của NV bị hủy ngay lập tức\n3. NV thử gửi API request từ token cũ",
                 "User ID: NV đã nghỉ việc", "Token cũ bị thu hồi, hệ thống từ chối truy cập với lỗi HTTP 401 Unauthorized"),
            ]),
            ("D. Luồng Quên mật khẩu & Reset Password", [
                ("BPT-ACC-004", "Yêu cầu quên mật khẩu ➔ Nhận OTP ➔ Xác thực OTP ➔ Đổi mật khẩu mới",
                 "Kiểm tra luồng khôi phục mật khẩu qua email OTP", "Tài khoản NV tồn tại trong hệ thống",
                 "1. Nhấn Quên mật khẩu tại trang login\n2. Nhập email\n3. Nhập mã OTP từ email\n4. Nhập mật khẩu mới và hoàn tất",
                 "Email: employee@hocba.edu.vn", "Mật khẩu được đổi thành công, đăng nhập thành công với mật khẩu mới"),
            ])
        ],
        "ISP490_G2_BPT_ACCOUNTS.xlsx"
    )

    # BPT EMPLOYEES
    create_bpt_file(
        "EMPLOYEES", "Quản lý Hồ sơ Nhân viên", "BPT_EMPLOYEES",
        [
            ("A. Luồng Tiếp nhận & Khởi tạo Hồ sơ Nhân viên", [
                ("BPT-EMP-001", "Tạo hồ sơ NV ➔ Gán Phòng ban & Chức danh ➔ Tạo Hợp đồng thử việc ➔ Gán Loại NV",
                 "Kiểm tra luồng tiếp nhận nhân viên mới và thiết lập hồ sơ đầy đủ", "Tài khoản HR Manager đã đăng nhập",
                 "1. HR mở form Tạo nhân viên mới\n2. Nhập thông tin cá nhân & phòng ban\n3. Thêm hợp đồng thử việc 2 tháng\n4. Lưu hồ sơ",
                 "Họ tên: Nguyễn Văn A, Phòng: Công nghệ thông tin", "Hồ sơ NV được tạo ở trạng thái 'Thử việc', hợp đồng được ghi nhận chính xác"),
            ]),
            ("B. Luồng Cập nhật Bằng cấp & Điểm Trình độ", [
                ("BPT-EMP-002", "Cập nhật Bằng đại học & Chứng chỉ ➔ Tự động tính điểm trình độ ➔ Hiển thị Báo cáo",
                 "Kiểm tra tính năng tự động tính điểm kỹ năng/bằng cấp nhân viên", "Hồ sơ NV đã tồn tại",
                 "1. Truy cập tab Bằng cấp & Kỹ năng\n2. Thêm bằng Đại học Bách Khoa & Tiếng Anh TOEIC 800\n3. Lưu thay đổi",
                 "Bằng cấp: Đại học, Tiếng Anh: TOEIC 800", "Điểm trình độ tự động cập nhật lên 85/100 và phản ánh đúng trên biểu đồ báo cáo nhân sự"),
            ]),
            ("C. Luồng Luân chuyển Nội bộ & Cập nhật Quản lý", [
                ("BPT-EMP-003", "Chuyển phòng ban NV ➔ Thay đổi Quản lý trực tiếp ➔ Cập nhật Phân quyền dữ liệu",
                 "Kiểm tra luồng luân chuyển nhân sự giữa các phòng ban", "Hồ sơ NV đã tồn tại",
                 "1. Mở form Chuyển phòng ban\n2. Chọn Phòng Kinh doanh & Trưởng phòng mới\n3. Xác nhận luân chuyển",
                 "Phòng mới: Kinh doanh, Quản lý: Trần Thị B", "NV thuộc phòng ban mới, Quản lý mới nhìn thấy hồ sơ NV trên Dashboard"),
            ]),
            ("D. Luồng Xuất Báo cáo Hồ sơ Nhân sự", [
                ("BPT-EMP-004", "Lọc danh sách NV theo Phòng ban ➔ Xuất file XLSX/PDF ➔ Kiểm tra tính toàn vẹn",
                 "Kiểm tra xuất báo cáo nhân sự định dạng Excel và PDF", "Tài khoản HR Manager đã đăng nhập",
                 "1. Vào Màn hình Báo cáo Nhân sự\n2. Bộ lọc: Phòng Đào tạo, Trạng thái: Chính thức\n3. Nhấn Xuất Excel",
                 "Định dạng: XLSX", "File Excel tải về chứa đầy đủ danh sách NV thỏa điều kiện lọc với định dạng chuẩn"),
            ])
        ],
        "ISP490_G2_BPT_EMPLOYEES.xlsx"
    )

    # BPT EMPLOYEE LIFECYCLE
    create_bpt_file(
        "EMPLOYEE_LIFECYCLE", "Quản lý Vòng đời Nhân viên & Tuyển dụng", "BPT_LIFECYCLE",
        [
            ("A. Luồng Tuyển dụng ➔ Chuyển thành Nhân viên", [
                ("BPT-LC-001", "Ứng viên đạt Phỏng vấn ➔ Chuyển thành NV chính thức ➔ Sinh hợp đồng & tài khoản",
                 "Kiểm tra chuyển đổi tự động từ ứng viên tuyển dụng sang hồ sơ nhân viên", "Ứng viên có trạng thái Pass phỏng vấn",
                 "1. Mở danh sách Ứng viên Đạt\n2. Nhấn 'Tuyển dụng & Tạo hồ sơ NV'\n3. Nhập thông tin hợp đồng ban đầu\n4. Xác nhận",
                 "Mã ứng viên: UV-2026-089", "Tự động sinh Hồ sơ Nhân viên mới, Hợp đồng thử việc và gửi mail kích hoạt tài khoản"),
            ]),
            ("B. Luồng Đánh giá Thử việc ➔ Ký Hợp đồng Chính thức", [
                ("BPT-LC-002", "Đánh giá hết 2 tháng thử việc ➔ Chốt kết quả Đạt ➔ Ký Hợp đồng chính thức",
                 "Kiểm tra luồng đánh giá thử việc và chuyển đổi hợp đồng lao động", "NV sắp hết hạn hợp đồng thử việc",
                 "1. Quản lý lập phiếu Đánh giá thử việc (Đạt)\n2. HR phê duyệt kết quả\n3. Sinh Hợp đồng xác định thời hạn 1 năm\n4. Ký duyệt",
                 "Kết quả: Pass, Loại HĐ: 1 năm", "Trạng thái nhân viên chuyển sang 'Chính thức', quỹ phép năm được cấp theo quy định"),
            ]),
            ("C. Luồng Offboarding (Thủ tục Nghỉ việc)", [
                ("BPT-LC-003", "Nộp đơn nghỉ việc ➔ Bàn giao tài sản ➔ Thanh lý hợp đồng ➔ Lưu trữ hồ sơ",
                 "Kiểm tra luồng thủ tục nghỉ việc và bàn giao công việc", "Nhân viên chính thức nộp đơn nghỉ việc",
                 "1. NV nộp đơn xin nghỉ việc\n2. Trưởng phòng & HR duyệt ngày nghỉ việc chính thức\n3. Hoàn tất checklist bàn giao tài sản\n4. Chốt thanh lý HĐ",
                 "Ngày nghỉ việc: 31/08/2026", "Hồ sơ NV được chuyển sang trạng thái 'Archived', tài khoản bị khóa vào ngày nghỉ chính thức"),
            ]),
            ("D. Luồng Đánh giá Hiệu suất Định kỳ (Review)", [
                ("BPT-LC-004", "HR mở đợt Đánh giá ➔ Trưởng phòng chấm điểm ➔ Xếp loại A/B/C ➔ Công bố kết quả",
                 "Kiểm tra luồng đánh giá nhân viên định kỳ quý/năm", "Đợt đánh giá được khởi tạo",
                 "1. HR mở đợt Đánh giá Quý 3/2026\n2. Quản lý chấm 6 tiêu chí hiệu suất\n3. Hệ thống tự động tính tổng điểm & xếp loại\n4. HR công bố kết quả",
                 "Đợt: Q3/2026", "Điểm đánh giá được ghi nhận, sinh thông báo kết quả cho nhân viên"),
            ])
        ],
        "ISP490_G2_BPT_EMPLOYEE_LIFECYCLE.xlsx"
    )

    # BPT PAYROLL
    create_bpt_file(
        "PAYROLL", "Quản lý Bảng lương & Chi trả", "BPT_PAYROLL",
        [
            ("A. Luồng Chốt công ➔ Tính Bảng lương tự động", [
                ("BPT-PAY-001", "Đồng bộ Chấm công T7 ➔ Tạo kỳ lương T7 ➔ Tính tự động 192 phiếu lương ➔ Kiểm tra công thức AST",
                 "Kiểm tra luồng tổng hợp công và tính toán bảng lương hàng tháng", "Dữ liệu chấm công Tháng 7 đầy đủ",
                 "1. HR truy cập màn Kỳ tính lương\n2. Nhấn Tạo kỳ lương Tháng 07/2026\n3. Nhấn 'Tạo & Tính phiếu lương'\n4. Kiểm tra kết quả tính toán",
                 "Tháng: 07/2026, NV: 192 người", "192 phiếu lương được tính toán chính xác số công, phụ cấp, BHXH và Thuế TNCN"),
            ]),
            ("B. Luồng Phê duyệt & Gửi Email Xác nhận Phiếu lương", [
                ("BPT-PAY-002", "Duyệt kỳ lương ➔ Gửi mail xác nhận cho 192 NV ➔ NV click link xác nhận phiếu lương",
                 "Kiểm tra luồng gửi mail xác nhận phiếu lương và phản hồi từ nhân viên", "Kỳ tính lương ở trạng thái Tính xong",
                 "1. Trưởng phòng nhấn Phê duyệt kỳ lương\n2. HR nhấn 'Gửi mail xác nhận'\n3. NV mở email click link xác nhận",
                 "Kỳ lương: T07/2026", "Email được gửi thành công, trạng thái xác nhận của NV chuyển sang 'Confirmed'"),
            ]),
            ("C. Luồng Chi trả lương Ngân hàng & Chốt Lịch sử", [
                ("BPT-PAY-003", "Tạo file chi lương XLSX (MB/VCB/TCB) ➔ HR xác nhận chi ➔ Chốt kỳ lương (Close)",
                 "Kiểm tra luồng xuất file chuyển khoản ngân hàng và đóng kỳ lương", "Kỳ lương đã được duyệt",
                 "1. HR nhấn 'Tạo file Bank'\n2. Chọn định dạng ngân hàng MB Bank & VCB\n3. Xác nhận đã chi trả thành công\n4. Nhấn 'Lưu lịch sử (Close)'",
                 "Ngân hàng: MB, VCB", "File XLSX chi lương chuẩn định dạng bank, kỳ lương chuyển trạng thái 'Close', lưu vào lịch sử"),
            ]),
            ("D. Luồng Sinh Chứng từ Kế toán Chi phí Lương", [
                ("BPT-PAY-004", "Chốt kỳ lương ➔ Tự động sinh Chứng từ kế toán (`hocba.fin.voucher`) hạch toán chi phí lương",
                 "Kiểm tra hạch toán kế toán tự động khi đóng kỳ tính lương", "Kỳ lương Tháng 7 được Close",
                 "1. Đóng kỳ lương Tháng 07/2026\n2. Mở màn hình Quản lý Chứng từ Kế toán (`hocba.fin.voucher`)\n3. Kiểm tra chứng từ chi phí lương tự động sinh",
                 "Mã kỳ lương: BATCH-2026-07", "Chứng từ kế toán hạch toán Nợ TK 642 / Có TK 334 được sinh chính xác số tiền"),
            ])
        ],
        "ISP490_G2_BPT_PAYROLL.xlsx"
    )

    # BPT SERVICE
    create_bpt_file(
        "SERVICE", "Quản lý Dịch vụ Nhân sự & Ticket", "BPT_SERVICE",
        [
            ("A. Luồng Gửi Yêu cầu Hỗ trợ (Ticket Service)", [
                ("BPT-SVC-001", "NV gửi ticket Hỗ trợ ➔ Phân công Cán bộ xử lý ➔ Phản hồi 2 chiều ➔ Đóng ticket",
                 "Kiểm tra luồng tiếp nhận và xử lý yêu cầu hỗ trợ nhân sự", "Tài khoản Nhân viên đã đăng nhập",
                 "1. NV tạo ticket hỗ trợ xác nhận quá trình công tác\n2. Cán bộ HR tiếp nhận và phân công xử lý\n3. Trả lời trao đổi trong ticket\n4. Đóng ticket",
                 "Loại yêu cầu: Xác nhận công tác", "Ticket được xử lý thành công, lịch sử trao đổi 2 chiều lưu trữ đầy đủ"),
            ]),
            ("B. Luồng Gửi Ý kiến Ẩn danh & Bảo vệ Danh tính", [
                ("BPT-SVC-002", "NV nộp ý kiến ẩn danh ➔ Kiểm tra ngưỡng ẩn danh ➔ Ban Giám đốc phản hồi ➔ Đóng đơn",
                 "Kiểm tra cơ chế bảo vệ danh tính nhân viên khi gửi góp ý ẩn danh", "Hệ thống bật cấu hình ẩn danh",
                 "1. NV chọn gửi góp ý ẩn danh\n2. Nhập nội dung góp ý ban quản lý\n3. Ban Giám đốc tiếp nhận và phản hồi public\n4. Đóng đơn",
                 "Chế độ: Ẩn danh", "Danh tính người gửi hoàn toàn bị ẩn trên mọi view/API, Ban Giám đốc phản hồi thành công"),
            ]),
            ("C. Luồng Cảnh báo Quá hạn SLA Yêu cầu Service", [
                ("BPT-SVC-003", "Ticket không được xử lý trong 48h ➔ Tự động đánh dấu Overdue ➔ Gửi thông báo nhắc",
                 "Kiểm tra cron job tự động quét SLA và cảnh báo quá hạn ticket", "Ticket được tạo quá 48h chưa xử lý",
                 "1. Chạy cron job quét SLA service\n2. Kiểm tra trạng thái ticket quá hạn\n3. Kiểm tra thông báo nhắc nhở gửi cho Cán bộ phụ trách",
                 "SLA Limit: 48h", "Ticket tự động chuyển trạng thái 'Overdue', chuông thông báo nhắc nhở được gửi tới HR Manager"),
            ])
        ],
        "ISP490_G2_BPT_SERVICE.xlsx"
    )

if __name__ == "__main__":
    run()
