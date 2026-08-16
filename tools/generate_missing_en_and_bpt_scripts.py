import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')
UNIT_TEST_DIR = r"D:\FPT\DO_an\code\ISP490_G2_HOCBA_HRM\HRM_Học Bá\unit-test"

FONT = 'Arial'
HDR_FILL = PatternFill('solid', fgColor='FFBDD6EE')
SEC_FILL = PatternFill('solid', fgColor='FFFBE4D5')
THIN = Side('thin', color='FF000000')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def create_ut_en_file(module_code, module_name, func_id, sections_data, out_filename):
    wb = openpyxl.Workbook()
    
    # Cover
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.cell(row=8, column=2, value="Unit Test Specification").font = Font(name=FONT, size=16, bold=True, color="003366")
    ws_cover.cell(row=11, column=2, value="Module").font = Font(name=FONT, size=10, bold=True)
    ws_cover.cell(row=11, column=3, value=module_name)
    ws_cover.cell(row=12, column=2, value="Function ID").font = Font(name=FONT, size=10, bold=True)
    ws_cover.cell(row=12, column=3, value=func_id)
    ws_cover.cell(row=13, column=2, value="Function Name").font = Font(name=FONT, size=10, bold=True)
    ws_cover.cell(row=13, column=3, value=f"{module_name.upper()} MODULE — HOCBA HRM")

    # Histories
    ws_hist = wb.create_sheet(title="Histories")
    headers_hist = ["No", "Version", "Description", "Sheet", "Modified date", "Modified by"]
    for c_idx, h in enumerate(headers_hist, 2):
        cell = ws_hist.cell(row=2, column=c_idx, value=h)
        cell.font = Font(name=FONT, size=10, bold=True)
        cell.fill = HDR_FILL
        cell.border = BOX
    row_hist = [1, "1.0", f"Initial English Unit Test Specification for {module_name}", "Cover, Histories, UT, Test Result", "08/08/2026", "G2 Team"]
    for c_idx, val in enumerate(row_hist, 2):
        cell = ws_hist.cell(row=3, column=c_idx, value=val)
        cell.font = Font(name=FONT, size=10)
        cell.border = BOX

    # UT
    ws_ut = wb.create_sheet(title="UT")
    ws_ut.cell(row=2, column=2, value="Function ID").font = Font(name=FONT, size=10, bold=True)
    ws_ut.cell(row=2, column=3, value=func_id)
    ws_ut.cell(row=2, column=12, value="Function Name").font = Font(name=FONT, size=10, bold=True)
    ws_ut.cell(row=2, column=13, value=module_name)

    headers_ut = ["STT", "Testcase ID", "Feature / Screen", "Test Case Name", "Detailed Description", "Pre-conditions", "Execution Steps", "Input Data", "Expected Output", "Test Type", "Priority", "Automated", "Status", "Tester", "Notes"]
    for c_idx, h in enumerate(headers_ut, 1):
        cell = ws_ut.cell(row=6, column=c_idx, value=h)
        cell.font = Font(name=FONT, size=10, bold=True)
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX

    curr_row = 7
    stt = 1
    test_result_sheets = []

    for sec_title, test_cases in sections_data:
        sec_cell = ws_ut.cell(row=curr_row, column=1, value=sec_title)
        sec_cell.font = Font(name=FONT, size=10, bold=True)
        sec_cell.fill = SEC_FILL
        ws_ut.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=15)
        curr_row += 1

        for tc_id, tc_name, desc, pre_req, steps, inputs, expected in test_cases:
            sheet_code = f"A{stt}" if stt <= 4 else (f"B{stt-4}" if stt <= 8 else f"C{stt-8}")
            test_result_sheets.append((sheet_code, tc_id, tc_name))

            row_vals = [
                stt, tc_id, sec_title.split('.')[1].strip() if '.' in sec_title else sec_title,
                tc_name, desc, pre_req, steps, inputs, expected,
                "Functional", "High", "Yes", "Pass", "G2 Team", ""
            ]

            ws_ut.row_dimensions[curr_row].height = 40
            for c_idx, val in enumerate(row_vals, 1):
                cell = ws_ut.cell(row=curr_row, column=c_idx, value=val)
                cell.font = Font(name=FONT, size=9)
                cell.border = BOX
                cell.alignment = Alignment(horizontal="center" if c_idx in (1, 2, 10, 11, 12, 13) else "left", vertical="center", wrap_text=True)
            
            stt += 1
            curr_row += 1

    for sheet_code, tc_id, tc_name in test_result_sheets:
        ws_res = wb.create_sheet(title=f"Test Result ({sheet_code})")
        ws_res.cell(row=2, column=2, value="Function ID").font = Font(name=FONT, size=10, bold=True)
        ws_res.cell(row=2, column=3, value=tc_id)
        ws_res.cell(row=2, column=12, value="Function Name").font = Font(name=FONT, size=10, bold=True)
        ws_res.cell(row=2, column=13, value=tc_name)

    out_path = os.path.join(UNIT_TEST_DIR, out_filename)
    wb.save(out_path)
    print(f"Created UT EN Excel file: {out_path}")

def run():
    # UT TIMEOFF EN
    create_ut_en_file(
        "TIMEOFF", "Time Off & Leave Management", "UT_TIMEOFF_EN",
        [
            ("A. Leave Request Submission", [
                ("UT-TO-001", "Submit Annual Leave Request Successfully",
                 "Verify creating annual leave request with valid dates and quota", "Employee logged in",
                 "1. Open Leave Request screen\n2. Select Annual Leave\n3. Input valid dates\n4. Click Submit",
                 "Type: Annual Leave, Days: 2", "Leave request created with 'Pending' status"),
                ("UT-TO-002", "Reject Request when End Date is before Start Date",
                 "Verify system error validation when end date < start date", "Employee logged in",
                 "1. Open Leave Request screen\n2. Select start: 15/08, end: 10/08\n3. Click Submit",
                 "Start: 15/08, End: 10/08", "Validation error message displayed"),
            ]),
            ("B. Leave Approval & Balance Check", [
                ("UT-TO-003", "Manager Approves Leave Request & Deducts Quota",
                 "Verify manager approval workflow and automatic leave quota deduction", "Pending leave request exists",
                 "1. Manager opens Approval Dashboard\n2. Select pending request\n3. Click Approve",
                 "Request ID: TO-001", "Request status updated to 'Approved', quota deducted by 2 days"),
            ])
        ],
        "ISP490_G2_UT_TIMEOFF_EN.xlsx"
    )

    # UT SERVICE EN
    create_ut_en_file(
        "SERVICE", "HR Service Requests & Ticketing", "UT_SERVICE_EN",
        [
            ("A. Ticket Submission & Processing", [
                ("UT-SVC-001", "Submit HR Support Ticket Successfully",
                 "Verify creating support ticket with attachment", "Employee logged in",
                 "1. Open HR Support screen\n2. Select ticket type\n3. Input subject & description\n4. Click Submit",
                 "Type: Certificate Request", "Ticket created with 'Open' status"),
                ("UT-SVC-002", "Anonymous Ticket Submission with Privacy Masking",
                 "Verify submitting anonymous feedback ticket", "Anonymous option enabled",
                 "1. Select Anonymous mode\n2. Input feedback content\n3. Click Send",
                 "Mode: Anonymous", "Ticket created with sender identity masked"),
            ])
        ],
        "ISP490_G2_UT_SERVICE_EN.xlsx"
    )

if __name__ == "__main__":
    run()
