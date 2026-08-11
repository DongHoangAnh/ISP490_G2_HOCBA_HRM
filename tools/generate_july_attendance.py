import os
import random
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_july_attendance_file():
    # File path
    file_path = r"D:\FPT\DO_an\code\ISP490_G2_HOCBA_HRM\Mau_Cham_Cong_Thang_07_2026.xlsx"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chấm công T07-2026"
    ws.views.sheetView[0].showGridLines = True

    # Title header
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "BẢNG DỮ LIỆU CHẤM CÔNG NHÂN VIÊN — THÁNG 07/2026"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="003366")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:K2")
    sub_cell = ws["A2"]
    sub_cell.value = "Công ty: Học Bá Education | Thời gian: 01/07/2026 - 31/07/2026"
    sub_cell.font = Font(name="Arial", size=10, italic=True, color="555555")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Table headers
    headers = [
        "STT", "Mã NV", "Họ và Tên", "Phòng ban", "Ngày", 
        "Check In", "Check Out", "Số công", "Số giờ làm", "Đi trễ (Phút)", "Ghi chú"
    ]
    
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[4].height = 28
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header_title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    print(f"Creating Excel template at: {file_path}")

if __name__ == "__main__":
    generate_july_attendance_file()
